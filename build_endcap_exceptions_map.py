"""
Endcap Exceptions Map
=====================
Maps the 1,884 endcap-program stores by set outcome — set successfully vs.
the four exception reasons from "Exceptions_Catalyst Endcap Stores.xlsx" —
against the backdrop of every other store carrying Catalyst (week 202627
Sales by Store). Writes a self-contained Leaflet map ->
endcap_exceptions_map.html.

"Set successfully" = endcap store with no exception filed.
Popups show wk202626 -> wk202627 15-lb Original units so the endcap lift is
visible per store.
"""
import csv
import json
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
GEO = os.path.join(HERE, "stores_geo.json")
OUTPUT = os.path.join(HERE, "endcap_exceptions_map.html")
EXC_FILE = os.path.join(HERE, "Exceptions_Catalyst Endcap Stores.xlsx")
PRIOR_WEEK, ENDCAP_WEEK = "202626", "202627"

SEGMENTS = {
    "set":       {"label": "Endcap set successfully", "color": "#1a9850"},
    "space":     {"label": "No available space", "color": "#d73027"},
    "inventory": {"label": "Not enough inventory", "color": "#f4a340"},
    "located":   {"label": "Product not located", "color": "#8e44ad"},
    "refusal":   {"label": "Store refusal", "color": "#252525"},
    "other":     {"label": "All other Catalyst stores", "color": "#b8c4cc"},
}
SHEET_TO_SEG = {
    "No Available Space": "space",
    "Not Enough Inventory": "inventory",
    "Product not Located": "located",
    "Store Refusal": "refusal",
}


def load_qty15(week):
    """store -> 15-lb Original POS qty for the given week."""
    wb = openpyxl.load_workbook(
        os.path.join(HERE, f"{week} Weekly Sales Report Catalyst.xlsx"),
        read_only=True)
    ws = wb["Sales by Store"]
    q15, zips, allq = {}, {}, defaultdict(float)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[1] is None:
            continue
        try:
            store = int(r[1])
        except (TypeError, ValueError):
            continue
        allq[store] += float(r[8] or 0)
        zips[store] = (str(r[5] or "")[:5], (r[4] or "").title(), r[3] or "")
        if str(r[0]).strip() == "CATALYST15ORIG":
            q15[store] = float(r[8] or 0)
    return q15, allq, zips


def main():
    # endcap roster (has its own lat/lon)
    endcap = {}
    with open(os.path.join(HERE, "endcap_stores.csv"), encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            endcap[int(r["store_number"])] = r

    # exceptions -> segment
    seg_of = {}
    wb = openpyxl.load_workbook(EXC_FILE, read_only=True)
    for sheet, seg in SHEET_TO_SEG.items():
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[4] is not None:
                seg_of[int(row[4])] = seg
    for s in endcap:
        seg_of.setdefault(s, "set")

    q15_prior, _, _ = load_qty15(PRIOR_WEEK)
    q15_now, all_now, zips = load_qty15(ENDCAP_WEEK)

    geo = json.load(open(GEO, encoding="utf-8"))
    stores, skipped = [], []
    for s, r in endcap.items():
        prior = ("none" if r["on_endcap_no_catalyst"] == "Y"
                 else f"{r['current_sku_count']} SKU" +
                      ("" if r["current_sku_count"] == "1" else "s") +
                      (", no 15-lb" if r["has_15O"] != "Y" else ""))
        if r["latitude"] and r["longitude"]:
            lat, lon = float(r["latitude"]), float(r["longitude"])
        else:  # fall back to the zip-code geocode cache
            g = geo.get(str(r["zip"])[:5]) or geo.get(zips.get(s, ("",))[0])
            if not g:
                skipped.append(s)
                continue
            lat, lon = g["lat"], g["lon"]
        stores.append({
            "store": s, "city": r["city"].title(), "state": r["state"],
            "lat": lat, "lon": lon,
            "seg": seg_of[s], "prior": prior,
            "q26": q15_prior.get(s, 0), "q27": q15_now.get(s, 0),
        })
    if skipped:
        print(f"[warn] {len(skipped)} endcap stores without coordinates skipped: {skipped}")
    # background: every other store selling Catalyst this week
    for s in all_now:
        if s in endcap:
            continue
        zip5, city, state = zips.get(s, ("", "", ""))
        g = geo.get(zip5)
        if not g:
            continue
        stores.append({
            "store": s, "city": city, "state": state,
            "lat": g["lat"], "lon": g["lon"], "seg": "other", "prior": None,
            "q26": q15_prior.get(s, 0), "q27": q15_now.get(s, 0),
        })

    counts = {k: sum(1 for x in stores if x["seg"] == k) for k in SEGMENTS}
    payload = {"week": ENDCAP_WEEK, "prior_week": PRIOR_WEEK,
               "segments": SEGMENTS, "counts": counts, "stores": stores,
               "n_endcap": len(endcap)}
    html = TEMPLATE.replace("/*DATA*/", json.dumps(payload, separators=(",", ":")))
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[ok] {len(stores)} stores mapped "
          f"({', '.join(f'{k}={v}' for k, v in counts.items())})")
    print(f"[ok] wrote {OUTPUT}")


TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Catalyst Endcap Exceptions Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  html, body { margin:0; height:100%; font-family:system-ui,Segoe UI,Arial,sans-serif; }
  #map { position:absolute; inset:0; }
  .hdr { position:absolute; top:10px; left:50px; right:10px; z-index:1000;
         background:#fff; border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.3);
         padding:8px 14px; display:inline-block; max-width:calc(100% - 80px); }
  .hdr h1 { margin:0; font-size:16px; }
  .hdr .sub { color:#666; font-size:12px; margin-top:2px; }
  .legend { position:absolute; bottom:20px; left:10px; z-index:1000; background:#fff;
            border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.3); padding:10px 14px;
            font-size:13px; }
  .legend .row { display:flex; align-items:center; gap:8px; margin:4px 0;
                 cursor:pointer; user-select:none; }
  .legend .row.off { opacity:.35; }
  .legend .dot { width:12px; height:12px; border-radius:50%; flex:0 0 auto; }
  .legend .n { color:#666; margin-left:auto; padding-left:12px; }
  .legend .hint { color:#999; font-size:11px; margin-top:6px; }
  .popup b { font-size:13px; }
  .popup table { border-collapse:collapse; margin-top:4px; font-size:12px; }
  .popup td { padding:1px 8px 1px 0; }
</style>
</head>
<body>
<div id="map"></div>
<div class="hdr">
  <h1>Catalyst Endcap — Set vs. Exceptions</h1>
  <div class="sub" id="sub"></div>
</div>
<div class="legend" id="legend"></div>
<script>
const DATA = /*DATA*/;
const map = L.map('map').setView([39.5, -96.5], 4);
L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',
  { attribution:'&copy; OpenStreetMap &copy; CARTO', maxZoom:18 }).addTo(map);

document.getElementById('sub').textContent =
  DATA.n_endcap.toLocaleString() + ' endcap stores \\u00b7 set W/E 08/08/26 \\u00b7 sales wk ' +
  DATA.prior_week + ' \\u2192 wk ' + DATA.week;

const groups = {};
for (const key in DATA.segments) groups[key] = L.layerGroup().addTo(map);

for (const s of DATA.stores) {
  const seg = DATA.segments[s.seg];
  const bg = s.seg === 'other';
  const m = L.circleMarker([s.lat, s.lon], {
    radius: bg ? 3 : 5, color:'#333', weight: bg ? 0 : .5,
    fillColor: seg.color, fillOpacity: bg ? .45 : .85 });
  let rows = '<tr><td>15-lb units wk' + DATA.prior_week + '</td><td>' + s.q26 + '</td></tr>' +
             '<tr><td>15-lb units wk' + DATA.week + '</td><td><b>' + s.q27 + '</b></td></tr>';
  if (s.prior !== null)
    rows += '<tr><td>Catalyst before endcap</td><td>' + s.prior + '</td></tr>';
  m.bindPopup('<div class="popup"><b>Store #' + s.store + '</b> \\u2014 ' +
    s.city + ', ' + s.state +
    '<br><span style="color:' + seg.color + ';font-weight:600">' + seg.label + '</span>' +
    '<table>' + rows + '</table></div>');
  m.addTo(groups[s.seg]);
}

const legend = document.getElementById('legend');
for (const key in DATA.segments) {
  const seg = DATA.segments[key], n = DATA.counts[key];
  const base = key === 'other' ? DATA.stores.length - DATA.n_endcap : DATA.n_endcap;
  const pct = key === 'other' ? '' : ' (' + Math.round(100 * n / DATA.n_endcap) + '%)';
  const row = document.createElement('div');
  row.className = 'row';
  row.innerHTML = '<span class="dot" style="background:' + seg.color + '"></span>' +
    seg.label + '<span class="n">' + n.toLocaleString() + pct + '</span>';
  row.onclick = () => {
    const off = row.classList.toggle('off');
    off ? map.removeLayer(groups[key]) : map.addLayer(groups[key]);
  };
  legend.appendChild(row);
}
const hint = document.createElement('div');
hint.className = 'hint';
hint.textContent = 'Percentages are of the 1,884 endcap stores \\u00b7 click a row to show/hide';
legend.appendChild(hint);
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
