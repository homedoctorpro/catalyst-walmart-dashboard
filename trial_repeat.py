"""
Trial & Repeat parser

Reads `Catalyst Trial and Repeat Report*.xlsx` (Walmart Luminate Trial & Repeat
module output) and produces a structured payload for the dashboard.

Structure:
{
  "meta": { "period_label", "submitted_by", "submitted_date", "sample_rate",
            "source_file", "weeks_count" },
  "weeks": ["2026-01-03", ...],         # ISO date strings (week-ending)
  "products": {
     "All Products" | "15O" | "34O" | "15U" | "34U": {
        "customers": { "1": [...], "2": [...], "3": [...], "4": [...], "5": [...], "6+": [...] },
        "dollars":   { same keys },
        "total_customers": [cum_total_per_week],
        "total_dollars":   [cum_total_per_week]
     }
  }
}
"""
from __future__ import annotations

import glob
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent

# Map by Walmart UPC suffix (more robust than name matching)
UPC_TO_SHORT = {
    "73550600020": "15O",
    "73550600022": "34O",
    "73550600021": "15U",
    "73550600023": "34U",
}

BUCKETS = ["1", "2", "3", "4", "5", "6+"]


def _latest_trial_repeat_file() -> Path | None:
    candidates = sorted(
        (p for p in glob.glob(str(ROOT / "Catalyst Trial and Repeat Report*.xlsx"))
         if not Path(p).name.startswith("~$")),          # skip Excel lock files
        key=lambda p: Path(p).stat().st_mtime,
        reverse=True,
    )
    return Path(candidates[0]) if candidates else None


def _product_key(item_name: str) -> str | None:
    if not item_name:
        return None
    name = str(item_name).strip()
    if name == "All Products":
        return "All Products"
    for upc, short in UPC_TO_SHORT.items():
        if upc in name:
            return short
    return None


def _bucket_key(label) -> str | None:
    if label is None:
        return None
    s = str(label).strip()
    if s.startswith("1 Transaction"):  return "1"
    if s.startswith("2 Transactions"): return "2"
    if s.startswith("3 Transactions"): return "3"
    if s.startswith("4 Transactions"): return "4"
    if s.startswith("5 Transactions"): return "5"
    if s.startswith("6+"):             return "6+"
    return None


def _parse_sheet(ws):
    """Return (weeks_iso, per_product_buckets) for one sheet.

    weeks_iso: list of ISO date strings
    per_product_buckets: { product_key: { bucket: [values_per_week] } }
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find header row — has 'Analysis Level' in col 0
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Analysis Level")
    header = rows[header_idx]
    # Columns 3..-1 are week-ending dates; last column is "Total"
    week_cols = []
    weeks_iso = []
    for ci, val in enumerate(header[3:], start=3):
        if isinstance(val, datetime):
            week_cols.append(ci)
            weeks_iso.append(val.strftime("%Y-%m-%d"))

    out: dict[str, dict[str, list[float]]] = {}
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None:
            continue
        # Skip the "Total" rows at the bottom (Analysis Level == 'Total')
        if r[0] == "Total":
            continue
        prod = _product_key(r[1])
        bucket = _bucket_key(r[2])
        if prod is None or bucket is None:
            continue
        vals = [float(r[ci] or 0) for ci in week_cols]
        out.setdefault(prod, {b: [0.0] * len(weeks_iso) for b in BUCKETS})
        out[prod][bucket] = vals

    return weeks_iso, out


def _parse_meta(wb) -> dict:
    meta: dict = {}
    if "Order Details" not in wb.sheetnames:
        return meta
    ws = wb["Order Details"]
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None:
            continue
        key, val = r[1], r[2] if len(r) > 2 else None
        if not key:
            continue
        if key == "Time Period":          meta["period_label"] = val
        elif key == "Submitted By":       meta["submitted_by"] = val
        elif key == "Submitted Date":     meta["submitted_date"] = val.strftime("%Y-%m-%d") if isinstance(val, datetime) else str(val) if val else None
        elif key == "Sample Rate":        meta["sample_rate"] = float(val) if val is not None else None
        elif key == "Calendar Type":      meta["calendar"] = val
        elif key == "Transaction Type":   meta["txn_type"] = val
    return meta


def _pack_buckets(buckets: dict, n_weeks: int) -> dict:
    """{bucket: [vals]} -> {"customers": ..., "total_customers": [...]}"""
    total = [sum(buckets[b][i] for b in BUCKETS) for i in range(n_weeks)]
    return {"customers": buckets, "total_customers": total}


def _build_ext(verbose: bool = True) -> dict | None:
    """Parse the lighter 'Weekly Repeats' chart export into an extended
    customer series (customers only — no $), which runs several weeks further
    than the full Trial & Repeat report. Two known formats:

    A. Full-report layout (e.g. `1612195_Trial & Repeat_WeeklyRepeats_*.xlsx`):
       a sheet with the 'Analysis Level' header — per-SKU + All Products.
       Extends ALL customer views; only $ stays on the full report's timeline.
    B. Legacy chart export (e.g. `1557501_Weekly_Repeats_trialrepeat.xlsx`,
       'Data' sheet): All-Products only — extends just the All-Products views.

    Returns { "weeks": [iso...], "customers": {bucket: [vals]},
              "total_customers": [vals], "source_file": str,
              "skus": {sku: {"customers":…, "total_customers":…}}  # format A only
            } or None.
    """
    candidates = {
        p for pat in ("*Weekly_Repeats*.xlsx", "*WeeklyRepeats*.xlsx")
        for p in glob.glob(str(ROOT / pat))
        if not Path(p).name.startswith("~$")             # skip Excel lock files
        and not Path(p).name.startswith("Catalyst Trial and Repeat Report")
    }
    if not candidates:
        return None
    src = Path(max(candidates, key=lambda p: Path(p).stat().st_mtime))
    wb = openpyxl.load_workbook(src, data_only=True)  # full (not read_only) — read_only mis-reports width

    # ── Format A: full-report layout ('Analysis Level' header) → per-SKU ──
    for name in wb.sheetnames:
        if name == "Order Details":
            continue
        ws = wb[name]
        if any(r and r[0] == "Analysis Level" for r in ws.iter_rows(values_only=True)):
            weeks_iso, prods = _parse_sheet(ws)
            if "All Products" not in prods:
                if verbose: print(f"  [trial_repeat] '{src.name}' sheet '{name}': no All Products row")
                return None
            out = {"weeks": weeks_iso, "source_file": src.name,
                   **_pack_buckets(prods["All Products"], len(weeks_iso)),
                   "skus": {k: _pack_buckets(v, len(weeks_iso))
                            for k, v in prods.items() if k != "All Products"}}
            return out

    # ── Format B: legacy 'Data' sheet — All-Products only ──
    if "Data" not in wb.sheetnames:
        if verbose: print(f"  [trial_repeat] '{src.name}' has no Data sheet - skipping ext")
        return None
    ws = wb["Data"]

    # Header row 1: col A label, then bucket columns in some order. Map by header text.
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    bucket_col = {}
    for label, col in header.items():
        b = _bucket_key(label)
        if b:
            bucket_col[b] = col
    if not bucket_col:
        if verbose: print(f"  [trial_repeat] '{src.name}' Data sheet: no transaction-bucket columns found")
        return None

    weeks_iso, cust = [], {b: [] for b in BUCKETS}
    for r in range(2, ws.max_row + 1):
        wk = ws.cell(row=r, column=1).value
        if wk in (None, ""):
            continue
        if isinstance(wk, datetime):
            weeks_iso.append(wk.strftime("%Y-%m-%d"))
        else:
            weeks_iso.append(datetime.strptime(str(wk).strip(), "%m/%d/%Y").strftime("%Y-%m-%d"))
        for b in BUCKETS:
            col = bucket_col.get(b)
            val = ws.cell(row=r, column=col).value if col else 0
            cust[b].append(float(val or 0))

    total = [sum(cust[b][i] for b in BUCKETS) for i in range(len(weeks_iso))]
    return {
        "weeks": weeks_iso,
        "customers": cust,
        "total_customers": total,
        "source_file": src.name,
    }


def build_trial_repeat_data(verbose: bool = True) -> dict | None:
    src = _latest_trial_repeat_file()
    if not src:
        if verbose: print("  [trial_repeat] No matching file found - skipping")
        return None

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    meta = _parse_meta(wb)
    meta["source_file"] = src.name

    if "WeeklyRepeats Transactions" not in wb.sheetnames:
        if verbose: print(f"  [trial_repeat] '{src.name}' missing WeeklyRepeats Transactions sheet")
        return None

    weeks_c, customers = _parse_sheet(wb["WeeklyRepeats Transactions"])
    weeks_d, dollars   = _parse_sheet(wb["WeeklyRepeats $$"]) if "WeeklyRepeats $$" in wb.sheetnames else (weeks_c, {})

    products: dict[str, dict] = {}
    for prod in sorted(set(customers) | set(dollars), key=lambda k: (k != "All Products", k)):
        cust_buckets = customers.get(prod, {b: [0.0] * len(weeks_c) for b in BUCKETS})
        doll_buckets = dollars.get(prod, {b: [0.0] * len(weeks_d) for b in BUCKETS})
        total_customers = [sum(cust_buckets[b][i] for b in BUCKETS) for i in range(len(weeks_c))]
        total_dollars   = [round(sum(doll_buckets[b][i] for b in BUCKETS), 2) for i in range(len(weeks_d))]
        products[prod] = {
            "customers": cust_buckets,
            "dollars":   doll_buckets,
            "total_customers": total_customers,
            "total_dollars":   total_dollars,
        }

    meta["weeks_count"] = len(weeks_c)
    meta["detail_through"] = weeks_c[-1] if weeks_c else None  # per-SKU & $ timeline end

    payload = {
        "meta": meta,
        "weeks": weeks_c,
        "products": products,
    }

    # Extend the customer series with the lighter weekly export (covers more
    # recent weeks). Only attach if it genuinely runs further.
    ext = _build_ext(verbose=verbose)
    if ext and len(ext["weeks"]) > len(weeks_c):
        skus_ext = ext.pop("skus", None)
        payload["allproducts_ext"] = ext
        if skus_ext:
            payload["skus_ext"] = skus_ext            # per-SKU extended customers
            meta["ext_has_sku"] = True
        meta["allproducts_through"] = ext["weeks"][-1]
        meta["allproducts_source"] = ext["source_file"]

    if verbose:
        last_total = products.get("All Products", {}).get("total_customers", [0])[-1]
        last_repeat = sum(products.get("All Products", {}).get("customers", {}).get(b, [0])[-1] for b in BUCKETS if b != "1")
        rate = (last_repeat / last_total * 100) if last_total else 0
        print(f"  [trial_repeat] {src.name}: {len(weeks_c)} weeks, "
              f"{len(products)} products, all-product repeat rate {rate:.1f}% (through {meta['detail_through']})")
        if "allproducts_ext" in payload:
            e = payload["allproducts_ext"]
            et = e["total_customers"][-1]
            er = sum(e["customers"][b][-1] for b in BUCKETS if b != "1")
            sku_note = f", per-SKU: {', '.join(payload['skus_ext'])}" if "skus_ext" in payload else ""
            print(f"  [trial_repeat]   + extended All-Products through {ext['weeks'][-1]} "
                  f"({len(e['weeks'])} weeks): repeat rate {er/et*100:.1f}% ({et:,.0f} customers){sku_note}")

    return payload


if __name__ == "__main__":
    import json
    data = build_trial_repeat_data(verbose=True)
    if data:
        out = ROOT / "trial_repeat_debug.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"Wrote {out.name}")
