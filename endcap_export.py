"""
Endcap Store Export — v2

Reads EndcapStoreList.xlsx (Allocations sheet) → 1,884 unique stores getting
the Catalyst Pet endcap. Cross-references against:
  1. Store & DC Addresses.xlsx → Wal-Mart Stores sheet (authoritative addresses)
  2. Latest weekly Sales by Store (current Catalyst SKU placements)
  3. stores_geocoded.json (precise lat/lon from prior Nominatim geocoding)

Geocodes any remaining stores by zip via pgeocode (caches in stores_geo.json).

Outputs:
  endcap_stores.csv   — flat table for data tools
  endcap_stores.xlsx  — same data as Excel
  endcap_map.html     — self-contained Leaflet map + sortable table tabs
"""
from __future__ import annotations

import csv
import glob
import json
from pathlib import Path

import openpyxl
import pgeocode

ROOT = Path(__file__).parent
ENDCAP_FILE = ROOT / "EndcapStoreList.xlsx"
ADDR_FILE = ROOT / "Store & DC Addresses.xlsx"
GEOCODED_FILE = ROOT / "Walmart Store List Map" / "data" / "processed" / "stores_geocoded.json"
ZIP_CACHE = ROOT / "stores_geo.json"

OUT_CSV = ROOT / "endcap_stores.csv"
OUT_XLSX = ROOT / "endcap_stores.xlsx"
OUT_MAP = ROOT / "endcap_map.html"

CATALYST_SKUS = ["CATALYST15ORIG", "CATALYST34LBORIGINAL", "CATALYST15UNSCEN", "CATALYSTPET34LBUNSCE"]
SKU_SHORT = {
    "CATALYST15ORIG": "15O",
    "CATALYST34LBORIGINAL": "34O",
    "CATALYST15UNSCEN": "15U",
    "CATALYSTPET34LBUNSCE": "34U",
}


def latest_weekly_file() -> Path | None:
    files = sorted(glob.glob(str(ROOT / "2026*.xlsx")))
    return Path(files[-1]) if files else None


def load_unique_endcap_stores() -> tuple[set[int], dict]:
    wb = openpyxl.load_workbook(ENDCAP_FILE, read_only=True)
    ws = wb["Allocations"]
    rows = list(ws.iter_rows(values_only=True))
    stores: set[int] = set()
    items_per_store: dict[int, set] = {}
    qty_per_store: dict[int, int] = {}
    for r in rows[1:]:
        if r[1] is None:
            continue
        s = int(r[1])
        stores.add(s)
        items_per_store.setdefault(s, set()).add(r[0])
        qty_per_store[s] = qty_per_store.get(s, 0) + int(r[7] or 0)
    return stores, {"items_per_store": items_per_store, "qty_per_store": qty_per_store}


def load_walmart_addresses() -> dict[int, dict]:
    """{store_num: {street, city, state, zip, store_name, type}} from authoritative source."""
    wb = openpyxl.load_workbook(ADDR_FILE, read_only=True)
    ws = wb["Wal-Mart Stores"]
    rows = list(ws.iter_rows(values_only=True))
    # Header is row index 3 (row 4 in 1-based)
    out: dict[int, dict] = {}
    for r in rows[4:]:
        if r[0] is None:
            continue
        try:
            s = int(r[0])
        except (TypeError, ValueError):
            continue
        out[s] = {
            "store_name": r[3],
            "type": r[4],
            "street": r[5],
            "city": r[6],
            "state": r[7],
            "zip": str(r[8]).split("-")[0].strip() if r[8] is not None else None,
        }
    return out


def load_current_skus() -> dict[int, set[str]]:
    """{store_num: {sku, ...}} from latest weekly Sales by Store sheet.

    A store-SKU pair appears in the sheet only if the store is currently traited
    for that item (Walmart's report excludes untraited combinations).
    """
    f = latest_weekly_file()
    if not f:
        return {}
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb["Sales by Store"]
    out: dict[int, set[str]] = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if r[1] is None or r[0] is None:
            continue
        try:
            s = int(r[1])
        except (TypeError, ValueError):
            continue
        sku = str(r[0]).strip().upper()
        if sku in CATALYST_SKUS:
            out.setdefault(s, set()).add(sku)
    return out


def load_geocoded_latlon() -> dict[int, tuple[float, float]]:
    with open(GEOCODED_FILE, encoding="utf-8") as f:
        data = json.load(f)
    out: dict[int, tuple[float, float]] = {}
    for s in data:
        if s.get("latitude") is not None and s.get("longitude") is not None:
            out[int(s["store_number"])] = (float(s["latitude"]), float(s["longitude"]))
    return out


def load_zip_cache() -> dict[str, dict]:
    if ZIP_CACHE.exists():
        with open(ZIP_CACHE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_zip_cache(cache: dict):
    with open(ZIP_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)


def geocode_zip(zip_code, nomi, cache):
    if not zip_code:
        return None, None
    z = str(zip_code).split("-")[0].strip().zfill(5)
    if z in cache and cache[z].get("lat") is not None:
        return cache[z]["lat"], cache[z]["lon"]
    res = nomi.query_postal_code(z)
    lat = res.latitude if res.latitude == res.latitude else None
    lon = res.longitude if res.longitude == res.longitude else None
    if lat is not None and lon is not None:
        cache[z] = {"lat": float(lat), "lon": float(lon)}
        return float(lat), float(lon)
    return None, None


def build_endcap_rows(verbose: bool = True) -> tuple[list[dict], dict]:
    """Build the per-store endcap row list. Returns (rows, summary)."""
    if not ENDCAP_FILE.exists():
        if verbose: print(f"  [endcap] {ENDCAP_FILE.name} not found - skipping")
        return [], {"total": 0, "mapped": 0, "addressed": 0, "no_catalyst": 0, "missing": []}

    if verbose: print("Loading endcap store list...")
    stores, meta = load_unique_endcap_stores()
    if verbose: print(f"  Unique stores: {len(stores)}")

    if verbose: print("Loading address sources...")
    addresses = load_walmart_addresses() if ADDR_FILE.exists() else {}
    if verbose: print(f"  Wal-Mart Stores master: {len(addresses)}")

    if verbose: print("Loading current Catalyst Pet SKU placements (latest week)...")
    current_skus = load_current_skus()
    if verbose: print(f"  Stores currently traited: {len(current_skus)}")

    if verbose: print("Loading prior geocodes...")
    geocoded_latlon = load_geocoded_latlon() if GEOCODED_FILE.exists() else {}
    if verbose: print(f"  Stores with precise lat/lon: {len(geocoded_latlon)}")

    zip_cache = load_zip_cache()
    nomi = pgeocode.Nominatim("us")

    rows: list[dict] = []
    geocoded_count = 0
    addressed = 0
    new_geocodes_before = len(zip_cache)
    no_catalyst_count = 0
    addr_missing: list[int] = []

    for s in sorted(stores):
        a = addresses.get(s, {})
        skus_now = current_skus.get(s, set())
        has = {sku: (sku in skus_now) for sku in CATALYST_SKUS}
        sku_count = sum(has.values())
        on_endcap_no_catalyst = sku_count == 0
        if on_endcap_no_catalyst:
            no_catalyst_count += 1

        rec = {
            "store_number": s,
            "store_name": a.get("store_name") or "",
            "store_type": a.get("type") or "",
            "street_address": a.get("street") or "",
            "city": a.get("city") or "",
            "state": a.get("state") or "",
            "zip": a.get("zip") or "",
            "has_15O": "Y" if has["CATALYST15ORIG"] else "",
            "has_34O": "Y" if has["CATALYST34LBORIGINAL"] else "",
            "has_15U": "Y" if has["CATALYST15UNSCEN"] else "",
            "has_34U": "Y" if has["CATALYSTPET34LBUNSCE"] else "",
            "current_sku_count": sku_count,
            "on_endcap_no_catalyst": "Y" if on_endcap_no_catalyst else "",
            "endcap_recommended_qty": meta["qty_per_store"].get(s, 0),
            "latitude": "",
            "longitude": "",
        }
        if rec["street_address"]:
            addressed += 1
        else:
            addr_missing.append(s)

        if s in geocoded_latlon:
            rec["latitude"], rec["longitude"] = geocoded_latlon[s]
            geocoded_count += 1
        elif rec["zip"]:
            lat, lon = geocode_zip(rec["zip"], nomi, zip_cache)
            if lat is not None:
                rec["latitude"] = lat
                rec["longitude"] = lon
                geocoded_count += 1

        rows.append(rec)

    if len(zip_cache) > new_geocodes_before:
        save_zip_cache(zip_cache)
        if verbose: print(f"  Geocoded {len(zip_cache) - new_geocodes_before} new zips (saved)")

    if verbose:
        print(f"\nResults:")
        print(f"  With address:                {addressed} / {len(rows)}")
        print(f"  With lat/lon:                {geocoded_count} / {len(rows)}")
        print(f"  On endcap, no Catalyst now:  {no_catalyst_count}")
        if addr_missing:
            print(f"  Still missing addresses:     {len(addr_missing)} -> {addr_missing[:10]}")

    summary = {
        "total": len(rows),
        "mapped": geocoded_count,
        "addressed": addressed,
        "no_catalyst": no_catalyst_count,
        "missing_addr_examples": addr_missing[:10],
    }
    return rows, summary


def main():
    rows, summary = build_endcap_rows(verbose=True)
    if not rows:
        return

    fieldnames = [
        "store_number", "store_name", "store_type",
        "street_address", "city", "state", "zip",
        "has_15O", "has_34O", "has_15U", "has_34U",
        "current_sku_count", "on_endcap_no_catalyst",
        "endcap_recommended_qty", "latitude", "longitude",
    ]

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"\nWrote {OUT_CSV.name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endcap Stores"
    ws.append(fieldnames)
    for r in rows:
        ws.append([r[k] for k in fieldnames])
    for col_idx, name in enumerate(fieldnames, start=1):
        max_len = max(len(str(name)), *(len(str(r[name])) for r in rows))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX.name}")

    map_payload = {
        "rows": rows,
        "total": summary["total"],
        "mapped": summary["mapped"],
        "no_catalyst": summary["no_catalyst"],
    }
    html = MAP_TEMPLATE.replace("/*DATA_PLACEHOLDER*/", json.dumps(map_payload, default=str))
    with open(OUT_MAP, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {OUT_MAP.name} ({summary['mapped']} stores plotted)")


MAP_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Catalyst Pet — Walmart Endcap Stores</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<style>
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; height: 100%; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; color: #222; }
  #header { background: #0057e7; color: #fff; padding: 14px 22px; box-shadow: 0 2px 6px rgba(0,0,0,.15); }
  #header h1 { margin: 0; font-size: 1.15rem; font-weight: 700; }
  #header .sub { font-size: 0.8rem; opacity: 0.9; margin-top: 4px; }
  .tabs { background: #fff; border-bottom: 1px solid #e0e0e0; padding: 0 22px; display: flex; gap: 4px; }
  .tab-btn { background: transparent; border: none; padding: 10px 18px; font-weight: 600; color: #666; cursor: pointer; border-bottom: 3px solid transparent; font-size: 0.9rem; }
  .tab-btn.active { color: #0057e7; border-bottom-color: #0057e7; }
  .tab-panel { display: none; }
  .tab-panel.active { display: block; }
  #map { height: calc(100vh - 110px); width: 100%; }
  .leaflet-popup-content { font-size: 0.85rem; line-height: 1.45; min-width: 200px; }
  .leaflet-popup-content b { color: #0057e7; }
  .leaflet-popup-content .badge { display: inline-block; padding: 1px 6px; border-radius: 4px; font-size: 0.72rem; font-weight: 700; margin-right: 3px; color: #fff; }
  .leaflet-popup-content .b-15O { background: #0057e7; }
  .leaflet-popup-content .b-34O { background: #e67300; }
  .leaflet-popup-content .b-15U { background: #00a86b; }
  .leaflet-popup-content .b-34U { background: #9c27b0; }
  .leaflet-popup-content .no-cat { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 0.72rem; font-weight: 700; background: #d32f2f; color: #fff; }

  .map-controls { background: #fff; padding: 10px 22px; border-bottom: 1px solid #e0e0e0; display: flex; gap: 14px; flex-wrap: wrap; align-items: center; font-size: 0.85rem; }
  .map-controls label { display: inline-flex; align-items: center; gap: 6px; cursor: pointer; user-select: none; }
  .legend-dot { display: inline-block; width: 12px; height: 12px; border-radius: 50%; border: 1px solid rgba(0,0,0,.2); vertical-align: middle; }
  .legend-dot.no-cat { background: #d32f2f; }
  .legend-dot.has-cat { background: #0057e7; }

  .table-wrap { padding: 14px 22px; }
  .table-controls { display: flex; gap: 12px; align-items: center; margin-bottom: 12px; flex-wrap: wrap; }
  .table-controls input[type=text] { padding: 7px 10px; border: 1px solid #ccc; border-radius: 6px; font-size: 0.9rem; min-width: 220px; }
  .table-controls label { font-size: 0.85rem; display: inline-flex; align-items: center; gap: 5px; }
  .table-controls .count { color: #666; font-size: 0.85rem; margin-left: auto; }
  table#stores { width: 100%; border-collapse: collapse; font-size: 0.82rem; background: #fff; box-shadow: 0 1px 4px rgba(0,0,0,.06); }
  table#stores th { position: sticky; top: 0; background: #f5f5f5; padding: 8px 10px; text-align: left; font-weight: 700; border-bottom: 2px solid #e0e0e0; cursor: pointer; white-space: nowrap; z-index: 5; }
  table#stores th:hover { background: #ececec; }
  table#stores td { padding: 6px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: middle; }
  table#stores td.center { text-align: center; }
  table#stores tr.no-cat { background: #fdecec; }
  table#stores tr:hover td { background: #f8f9ff; }
  table#stores tr.no-cat:hover td { background: #fbdedb; }
  .sku-chip { display: inline-block; padding: 2px 7px; border-radius: 4px; font-size: 0.72rem; font-weight: 700; color: #fff; margin-right: 2px; }
  .sku-chip.s-15O { background: #0057e7; }
  .sku-chip.s-34O { background: #e67300; }
  .sku-chip.s-15U { background: #00a86b; }
  .sku-chip.s-34U { background: #9c27b0; }
  .no-cat-flag { color: #d32f2f; font-weight: 700; }
</style>
</head>
<body>
<div id="header">
  <h1>Catalyst Pet — Walmart Endcap Stores</h1>
  <div class="sub" id="sub"></div>
</div>
<div class="tabs">
  <button class="tab-btn active" onclick="switchTab('map', this)">Map</button>
  <button class="tab-btn" onclick="switchTab('table', this)">Table</button>
</div>

<div id="tab-map" class="tab-panel active">
  <div class="map-controls">
    <label><input type="checkbox" id="show-has" checked> <span class="legend-dot has-cat"></span> Stocks Catalyst</label>
    <label><input type="checkbox" id="show-no" checked> <span class="legend-dot no-cat"></span> On endcap, no Catalyst yet</label>
  </div>
  <div id="map"></div>
</div>

<div id="tab-table" class="tab-panel">
  <div class="table-wrap">
    <div class="table-controls">
      <input type="text" id="search" placeholder="Search store #, city, state, address…">
      <label><input type="checkbox" id="only-no-cat"> Only stores without Catalyst</label>
      <span class="count" id="row-count"></span>
    </div>
    <div style="overflow-x:auto">
    <table id="stores">
      <thead><tr>
        <th data-col="store_number">Store #</th>
        <th data-col="street_address">Address</th>
        <th data-col="city">City</th>
        <th data-col="state">State</th>
        <th data-col="zip">Zip</th>
        <th data-col="has_15O" class="center">15O</th>
        <th data-col="has_34O" class="center">34O</th>
        <th data-col="has_15U" class="center">15U</th>
        <th data-col="has_34U" class="center">34U</th>
        <th data-col="current_sku_count" class="center">Current SKUs</th>
        <th data-col="endcap_recommended_qty" class="center">Endcap Qty</th>
      </tr></thead>
      <tbody id="tbody"></tbody>
    </table>
    </div>
  </div>
</div>

<script>
const DATA = /*DATA_PLACEHOLDER*/;
const ROWS = DATA.rows;
const SKU_LABELS = { has_15O: "15O", has_34O: "34O", has_15U: "15U", has_34U: "34U" };

document.getElementById("sub").textContent =
  `${DATA.total.toLocaleString()} unique endcap stores · `
  + `${DATA.mapped.toLocaleString()} plotted · `
  + `${DATA.no_catalyst.toLocaleString()} on endcap but not currently stocking Catalyst`;

function switchTab(name, btn) {
  document.querySelectorAll(".tab-btn").forEach(b => b.classList.remove("active"));
  document.querySelectorAll(".tab-panel").forEach(p => p.classList.remove("active"));
  btn.classList.add("active");
  document.getElementById("tab-" + name).classList.add("active");
  if (name === "map" && window._map) setTimeout(() => window._map.invalidateSize(), 50);
}

// ─── MAP ───────────────────────────────────────────────────────────────────
const map = L.map("map", { preferCanvas: true }).setView([39.5, -98.35], 5);
window._map = map;
L.tileLayer("https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png", {
  attribution: "© OpenStreetMap",
  maxZoom: 19,
}).addTo(map);

const clusterHas = L.markerClusterGroup({
  chunkedLoading: true, showCoverageOnHover: false, maxClusterRadius: 45,
  iconCreateFunction: c => L.divIcon({
    html: `<div style="background:#0057e7;color:#fff;border-radius:50%;width:32px;height:32px;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:0.75rem;border:2px solid #fff;box-shadow:0 1px 3px rgba(0,0,0,.3)">${c.getChildCount()}</div>`,
    className: "", iconSize: [32, 32],
  }),
});
const clusterNo = L.markerClusterGroup({
  chunkedLoading: true, showCoverageOnHover: false, maxClusterRadius: 45,
  iconCreateFunction: c => L.divIcon({
    html: `<div style="background:#d32f2f;color:#fff;border-radius:50%;width:32px;height:32px;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:0.75rem;border:2px solid #fff;box-shadow:0 1px 3px rgba(0,0,0,.3)">${c.getChildCount()}</div>`,
    className: "", iconSize: [32, 32],
  }),
});

function makePopup(r) {
  const addr = [r.street_address, r.city, r.state, r.zip].filter(Boolean).join(", ");
  const skus = ["15O","34O","15U","34U"].filter(k => r["has_" + k] === "Y");
  const skuHtml = skus.length
    ? skus.map(k => `<span class="badge b-${k}">${k}</span>`).join("")
    : `<span class="no-cat">⚠ No Catalyst SKUs yet</span>`;
  return `<b>Store #${r.store_number}</b>${r.store_name ? " · " + r.store_name : ""}<br>`
       + `${addr || "<i>no address</i>"}<br>`
       + `<div style="margin-top:6px"><strong>Currently stocked:</strong> ${skuHtml}</div>`
       + `<div style="margin-top:4px;color:#666;font-size:0.78rem">Endcap recommended qty: ${r.endcap_recommended_qty}</div>`;
}

ROWS.forEach(r => {
  if (r.latitude === "" || r.longitude === "") return;
  const noCat = r.on_endcap_no_catalyst === "Y";
  const color = noCat ? "#d32f2f" : "#0057e7";
  const marker = L.circleMarker([r.latitude, r.longitude], {
    radius: noCat ? 7 : 5,
    color: color, fillColor: color,
    fillOpacity: noCat ? 0.85 : 0.65,
    weight: noCat ? 2 : 1,
  });
  marker.bindPopup(makePopup(r));
  (noCat ? clusterNo : clusterHas).addLayer(marker);
});
map.addLayer(clusterHas);
map.addLayer(clusterNo);

document.getElementById("show-has").addEventListener("change", e => {
  if (e.target.checked) map.addLayer(clusterHas); else map.removeLayer(clusterHas);
});
document.getElementById("show-no").addEventListener("change", e => {
  if (e.target.checked) map.addLayer(clusterNo); else map.removeLayer(clusterNo);
});

// ─── TABLE ─────────────────────────────────────────────────────────────────
let sortCol = "store_number";
let sortAsc = true;

function chip(v, key) {
  return v === "Y" ? `<span class="sku-chip s-${key}">${key}</span>` : "";
}

function renderTable() {
  const q = document.getElementById("search").value.trim().toLowerCase();
  const onlyNo = document.getElementById("only-no-cat").checked;
  let rows = ROWS.slice();
  if (onlyNo) rows = rows.filter(r => r.on_endcap_no_catalyst === "Y");
  if (q) {
    rows = rows.filter(r =>
      String(r.store_number).includes(q) ||
      (r.street_address && r.street_address.toLowerCase().includes(q)) ||
      (r.city && r.city.toLowerCase().includes(q)) ||
      (r.state && r.state.toLowerCase().includes(q)) ||
      (r.zip && r.zip.includes(q))
    );
  }
  rows.sort((a, b) => {
    const av = a[sortCol], bv = b[sortCol];
    const an = typeof av === "number" ? av : (parseFloat(av) || av || "");
    const bn = typeof bv === "number" ? bv : (parseFloat(bv) || bv || "");
    if (an < bn) return sortAsc ? -1 : 1;
    if (an > bn) return sortAsc ? 1 : -1;
    return 0;
  });

  const tbody = document.getElementById("tbody");
  tbody.innerHTML = rows.map(r => `
    <tr class="${r.on_endcap_no_catalyst === "Y" ? "no-cat" : ""}">
      <td>${r.store_number}</td>
      <td>${r.street_address || ""}</td>
      <td>${r.city || ""}</td>
      <td>${r.state || ""}</td>
      <td>${r.zip || ""}</td>
      <td class="center">${chip(r.has_15O, "15O")}</td>
      <td class="center">${chip(r.has_34O, "34O")}</td>
      <td class="center">${chip(r.has_15U, "15U")}</td>
      <td class="center">${chip(r.has_34U, "34U")}</td>
      <td class="center">${r.current_sku_count === 0 ? '<span class="no-cat-flag">0</span>' : r.current_sku_count}</td>
      <td class="center">${r.endcap_recommended_qty}</td>
    </tr>
  `).join("");
  document.getElementById("row-count").textContent = `${rows.length.toLocaleString()} of ${ROWS.length.toLocaleString()} stores`;
}

document.querySelectorAll("#stores th").forEach(th => {
  th.addEventListener("click", () => {
    const col = th.dataset.col;
    if (sortCol === col) sortAsc = !sortAsc;
    else { sortCol = col; sortAsc = true; }
    renderTable();
  });
});
document.getElementById("search").addEventListener("input", renderTable);
document.getElementById("only-no-cat").addEventListener("change", renderTable);
renderTable();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
