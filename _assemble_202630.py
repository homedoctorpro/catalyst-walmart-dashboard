#!/usr/bin/env python3
"""Assemble '202630 Weekly Sales Report Catalyst.xlsx' in Hailey's layout.

Week 202630: Hailey was out and Jeff (StoR) sent a master workbook
('Walmart Data 8-31-2026.xlsx') whose Sales-by-Store tab was accidentally
re-run for week 202629. Sources used here instead:

  - In-store metrics : Jeff's 'LW POS Sales' tab (genuine 202630, keyed by
                       Prime Item Nbr; different column layout than Hailey's)
  - Ecomm LW         : Jeff's 'Catalyst E comm Only LW Sales' tab (verbatim)
  - Sales by Store   : our own Scintilla Report Builder pull for WMT last week
                       ('Catalyst Weekly Sales by Store' report, 4 SKUs)

Output sheets use Hailey's names/positions so extract_data.py's detect_sheets
and position-based parsers work unchanged.

Usage: python _assemble_202630.py <scintilla_download.xlsx>
"""
import sys
import openpyxl
from openpyxl import Workbook

JEFF = "_jeff_202630.xlsx"
OUT  = "202630 Weekly Sales Report Catalyst.xlsx"
WEEK = "202630"

ITEMS = {  # prime item nbr -> canonical item name (order matters: Hailey's row order)
    680268871: "CATALYST15ORIG",
    680065761: "CATALYST34LBORIGINAL",
    680268872: "CATALYST15UNSCEN",
    680065800: "CATALYSTPET34LBUNSCE",
}

# Jeff's 'LW POS Sales' LW-block columns (0-based)
J_POS, J_QTY, J_INSTOCK, J_USW, J_MD = 4, 8, 12, 36, 44
J_ITEMCOL = 3  # Prime Item Nbr

# Hailey-format instore positions (0-based): parser reads rows 2-5 SKUs + row 6 Total
H_POS, H_QTY, H_INSTOCK, H_USW, H_MD = 1, 7, 9, 33, 41
H_WIDTH = 44


def load_jeff_instore(wb):
    ws = wb["LW POS Sales"]
    rows = list(ws.iter_rows(values_only=True))
    by_item, total = {}, None
    item_rows = [r for r in rows[2:] if r[J_ITEMCOL] in ITEMS]
    for r in item_rows:
        by_item[r[J_ITEMCOL]] = r
    if len(by_item) != 4:
        raise SystemExit(f"expected 4 catalyst item rows, found {len(by_item)}")
    want = sum(r[J_POS] or 0 for r in by_item.values())
    for r in rows[2:]:
        if r[J_ITEMCOL] == "Total" and r[J_POS] and abs(r[J_POS] - want) < 1.0:
            total = r
            break
    if total is None:
        raise SystemExit("catalyst subtotal row not found in LW POS Sales")
    return by_item, total


def write_instore(out_wb, by_item, total, bysums):
    """bysums: {item_name: (pos_dollars, pos_qty)} summed from the by-store pull.

    POS $/qty come from the by-store data so the instore totals and the
    Sales-by-Store sheet agree exactly (as Hailey's workbooks always do, and
    matching the numbers quoted in Jeff's email). Ratio metrics (instock %,
    U/S/W, markdown %) come from Jeff's LW POS Sales tab — both are 202630,
    ~0.15% apart due to pull timing.
    """
    ws = out_wb.create_sheet("CATALYST- LW Sales")
    meta = [""] * H_WIDTH
    meta[0] = "Time Range Name"
    for c in (H_POS, H_QTY, H_INSTOCK, H_USW, H_MD):
        meta[c] = "LW"
    hdr = [""] * H_WIDTH
    hdr[0] = "All Links Item Desc"
    hdr[H_POS], hdr[H_QTY] = "POS $ TY", "POS Qty TY"
    hdr[H_INSTOCK], hdr[H_USW], hdr[H_MD] = "Instock % TY", "U/S/W TY", "Markdown % Sales TY"
    ws.append(meta)
    ws.append(hdr)

    def metric_row(name, src, pos, qty):
        row = [""] * H_WIDTH
        row[0] = name
        row[H_POS], row[H_QTY] = pos, qty
        row[H_INSTOCK], row[H_USW], row[H_MD] = src[J_INSTOCK], src[J_USW], src[J_MD]
        return row

    for nbr, name in ITEMS.items():
        pos, qty = bysums[name]
        ws.append(metric_row(name, by_item[nbr], pos, qty))
    tot_pos = round(sum(v[0] for v in bysums.values()), 2)
    tot_qty = sum(v[1] for v in bysums.values())
    ws.append(metric_row("Total", total, tot_pos, tot_qty))


def write_ecomm(out_wb, jeff_wb):
    src = jeff_wb["Catalyst E comm Only LW Sales"]
    ws = out_wb.create_sheet("CATALYST LW Ecomm")
    for r in src.iter_rows(values_only=True):
        ws.append(list(r))


def write_bystore(out_wb, scintilla_path):
    wb = openpyxl.load_workbook(scintilla_path, read_only=True, data_only=True)
    ws_in = wb["Sheet1"]
    rows = list(ws_in.iter_rows(values_only=True))
    hdr = [str(h or "") for h in rows[0]]

    def col(sub):
        for i, h in enumerate(hdr):
            if sub in h:
                return i
        raise SystemExit(f"column {sub!r} missing from scintilla export {hdr}")

    c = {
        "week":     col("walmart_calendar_week"),
        "desc":     col("prime_item_description"),
        "store":    col("store_number"),
        "street":   col("street_address_line_1"),
        "state":    col("state_or_province_name"),
        "city":     col("city_name"),
        "zip":      col("zip_code_or_postal_code"),
        "dc":       col("distribution_center_number"),
        "pos":      col("pos_sales_this_year"),
        "qty":      col("pos_quantity_this_year"),
        "onhand":   col("store_on_hand_quantity_this_year"),
        "transit":  col("store_in_transit_quantity_this_year"),
        "onorder":  col("store_on_order_quantity_this_year"),
        "traited":  col("traited_store_count_this_year"),
    }
    weeks = {r[c["week"]] for r in rows[1:] if r[c["week"]] is not None}
    if weeks != {int(WEEK)}:
        raise SystemExit(f"scintilla export covers weeks {sorted(weeks)}, expected {WEEK}")

    bysums = {}
    for r in rows[1:]:
        if r[c["store"]] is None:
            continue
        name = str(r[c["desc"]]).strip()
        s, q = bysums.get(name, (0.0, 0))
        bysums[name] = (s + float(r[c["pos"]] or 0), q + int(r[c["qty"]] or 0))
    bysums = {k: (round(v[0], 2), v[1]) for k, v in bysums.items()}

    tr = f"Time Range 1\nMIN:{WEEK}\nMAX:{WEEK}\n"
    ws = out_wb.create_sheet("CATALYST Sales by Store")
    ws.append([
        "item_name", "store_number", "street_address_line_1",
        "state_or_province_name", "city_name", "zip_code_or_postal_code",
        "distribution_center_number",
        tr + "POS Sales", tr + "POS Quantity", tr + "On Hand Quantity",
        tr + "In Transit Quantity", tr + "On Order Quantity",
        tr + "Total Pipeline Quantity", tr + "Traited Store Count",
        tr + "U/S/W Traited",
    ])
    n = 0
    for r in rows[1:]:
        if r[c["store"]] is None:
            continue
        onhand  = int(r[c["onhand"]] or 0)
        transit = int(r[c["transit"]] or 0)
        onorder = int(r[c["onorder"]] or 0)
        qty     = int(r[c["qty"]] or 0)
        traited = int(r[c["traited"]] or 0)
        ws.append([
            str(r[c["desc"]]).strip(), int(r[c["store"]]),
            r[c["street"]], r[c["state"]], r[c["city"]], str(r[c["zip"]]),
            r[c["dc"]],
            float(r[c["pos"]] or 0), qty, onhand,
            transit, onorder, onhand + transit + onorder,
            traited, (qty if traited else None),
        ])
        n += 1
    return n, bysums


def main():
    if len(sys.argv) != 2:
        raise SystemExit("usage: python _assemble_202630.py <scintilla_download.xlsx>")
    jeff_wb = openpyxl.load_workbook(JEFF, read_only=True, data_only=True)
    by_item, total = load_jeff_instore(jeff_wb)

    out = Workbook()
    out.remove(out.active)
    nrows, bysums = write_bystore(out, sys.argv[1])
    write_instore(out, by_item, total, bysums)
    write_ecomm(out, jeff_wb)
    out.save(OUT)

    tot_pos = sum(v[0] for v in bysums.values())
    tot_qty = sum(v[1] for v in bysums.values())
    print(f"[OK] wrote {OUT}")
    print(f"  instore/bystore total: ${tot_pos:,.2f} / {tot_qty:,} units")
    print(f"  jeff LW POS Sales tab: ${total[J_POS]:,.2f} / {total[J_QTY]:,} units")
    print(f"  bystore rows : {nrows}")
    for k, v in sorted(bysums.items()):
        print(f"    {k:22s} ${v[0]:>12,.2f} {v[1]:>7,}")


if __name__ == "__main__":
    main()
