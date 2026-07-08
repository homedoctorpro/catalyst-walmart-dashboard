"""
Chewy Sales Dashboard ETL
=========================
Reads the Chewy "Catalyst_Feline Fresh SKU L52W Report" Excel, builds a JSON
payload, and injects it into chewy_dashboard_template.html -> chewy_dashboard.html.

Data sources:
  - Sheet "FY25_FY26 YTD Data" : flat monthly table (16 mo), source of the time
    series -> Units, Retail $ (Net Sales), Autoship %, PDP OOS%  (per Chewy SKU).
  - Sheet "Summary"            : Chewy's own FY26-YTD vs FY25-YTD retail rollup
    (authoritative), plus the internal Lignetics SKU codes (CT01.. / FW01).
  - Brand Snapshot PDFs        : avg customer rating + top-10 states (latest month).
    These are the only source of brand-level wholesale ("Chewy's Historical
    Purchases"); we keep them as reference only. Per-SKU wholesale $ is computed
    as WHOLESALE_PRICES[part] * units once prices are supplied.

Wholesale prices: fill WHOLESALE_PRICES below (keyed by Chewy part number) and
re-run. Until then wholesale $ renders as pending.
"""
import json
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Chewy Brand Snapshots",
                    "Catalyst_Feline Fresh SKU L52W Report 5.4 (1).xlsx")
TEMPLATE = os.path.join(HERE, "chewy_dashboard_template.html")
OUTPUT = os.path.join(HERE, "chewy_dashboard.html")

# ---------------------------------------------------------------------------
# Wholesale unit prices, keyed by Chewy part number.  wholesale $ = price *
# units_sold.  Each entry: {"base": <through Sep 2025>, "new": <from Oct 2025>}.
# Total changeover to "new" pricing at WHOLESALE_CHANGEOVER (Sep is a blended
# month but uses Aug/base pricing per Lignetics).  None = no price -> pending.
# Prices per Lignetics wholesale sheet (Catalyst only; FF prices TBD).
# ---------------------------------------------------------------------------
WHOLESALE_CHANGEOVER = "2025-10"  # first month billed at "new" pricing
WHOLESALE_PRICES = {
    "241757": {"base": 9.80, "new": 10.29},    # CT01 Healthy 10-lb
    "241758": {"base": 15.75, "new": 16.54},   # CT02 Healthy 20-lb
    "965502": {"base": 23.25, "new": 24.37},   # CT08 Healthy 30-lb
    "241760": {"base": 9.80, "new": 10.29},    # CT03 Unscented 10-lb
    "241761": {"base": 15.75, "new": 16.49},   # CT04 Unscented 20-lb
    "241763": {"base": 9.80, "new": 10.29},    # CT05 Multi-Cat 10-lb
    "241764": {"base": 15.75, "new": 16.52},   # CT06 Multi-Cat 20-lb
    "1633142": {"base": 13.29, "new": 13.95},  # CT19 Pine Pellet 20-lb
    "1665670": {"base": 24.49, "new": 24.49},  # CT09 Sisal Mat
    "1674830": {"base": 16.79, "new": 16.79},  # FW01 Litter Scoop
    "1685430": {"base": 6.19, "new": 6.19},    # CT18 Poop Bags
    "1932182": {"base": 6.49, "new": 6.49},    # Feline Fresh Pine 10-lb
    "1932190": {"base": 9.79, "new": 9.79},    # Feline Fresh Pine 20-lb
    "1932198": {"base": 15.99, "new": 15.99},  # Feline Fresh Pine 40-lb
}


def wholesale_price(part, month):
    """Per-unit wholesale price for a part in a given 'YYYY-MM' month."""
    p = WHOLESALE_PRICES.get(part)
    if not p:
        return None
    return p["new"] if month >= WHOLESALE_CHANGEOVER else p["base"]

# Per-SKU customer sales UNITS for the two most recent months, read from the
# May-2026 and June-2026 Brand Snapshot PDFs (Excel only had a partial May).
# Verified: PDF "Customer Sales Units" == Excel "Units Sold" (exact monthly
# match Jul'25-Apr'26). Values are the PDF "Top 10 Products" lists; Poop Bags
# (11th, not in top 10) is the brand total minus the top-10 sum.
PDF_MONTHLY_UNITS = {
    "2026-05": {"241764": 7775, "241761": 1310, "241758": 1255, "241763": 1175,
                "241757": 453, "965502": 369, "241760": 304, "1633142": 143,
                "1665670": 72, "1674830": 48, "1685430": 5,
                "1932190": 484, "1932198": 248, "1932182": 203},
    "2026-06": {"241764": 7918, "241761": 1377, "241758": 1366, "241763": 1037,
                "241757": 373, "965502": 325, "241760": 253, "1633142": 159,
                "1665670": 45, "1674830": 31, "1685430": 11,
                "1932198": 365, "1932190": 333, "1932182": 183},
}

# Shelf (list) retail prices -> "implied retail $" = units * shelf price, shown
# alongside actual Net Sales. Catalyst LITTER steps +5% at RETAIL_CHANGEOVER
# (Oct 2025); accessories (mat/scoop/poop) and Feline Fresh do NOT step.
RETAIL_CHANGEOVER = "2025-10"
SHELF_PRICES = {
    "241757": 14.99, "241760": 14.99, "241763": 14.99,   # Catalyst 10-lb
    "241758": 24.99, "241761": 24.99, "241764": 24.99,    # Catalyst 20-lb
    "1633142": 24.99,                                     # Pine 20-lb
    "965502": 34.99,                                      # Healthy 30-lb
    "1665670": 34.99,                                     # Sisal Mat
    "1674830": 23.99,                                     # Litter Scoop
    "1685430": 9.99,                                      # Poop Bags
    "1932182": 9.99, "1932190": 14.99, "1932198": 23.49,  # Feline Fresh
}
# Catalyst litter only -> the +5% Oct 2025 step applies here (retail & wholesale).
STEP_PARTS = {"241757", "241760", "241763", "241758", "241761", "241764",
              "1633142", "965502"}
PRICE_REF_MONTHS = ["2025-11", "2025-12", "2026-01", "2026-02", "2026-03",
                    "2026-04"]


def shelf_price(part, month):
    """List retail $/unit; Catalyst litter steps +5% from Oct 2025."""
    p = SHELF_PRICES.get(part)
    if p and part in STEP_PARTS and month >= RETAIL_CHANGEOVER:
        p = round(p * 1.05, 2)
    return p


# Chewy Canada Feline Fresh (kg SKUs) — reported in the Aug'25-Mar'26 Brand
# Snapshots only (dropped from Apr 2026 on; never in the Excel; NOT in the US
# brand totals). 18.14-kg = 40-lb, 9.07-kg = 20-lb. Small volume.
CANADA_FF = {
    "40-lb (18.14 kg)": {"2025-08": 50, "2025-12": 56, "2026-01": 49,
                         "2026-02": 54, "2026-03": 38},
    "20-lb (9.07 kg)":  {"2025-08": 31, "2025-12": 11, "2026-01": 40,
                         "2026-02": 27, "2026-03": 19},
}
CANADA_WHOLESALE = {"40-lb (18.14 kg)": 15.99, "20-lb (9.07 kg)": 9.79}

# Latest-month brand snapshot (from the June 2026 Brand Snapshot PDFs).
SNAPSHOT = {
    "Catalyst": {
        "month": "June 2026",
        "rating": 4.19,
        "top_states": [["CA", 1293], ["PA", 762], ["NY", 761], ["TX", 689],
                       ["FL", 679], ["OH", 578], ["NC", 457], ["MI", 450],
                       ["IL", 414], ["VA", 404]],
    },
    "Feline Fresh": {
        "month": "June 2026",
        "rating": 4.02,
        "top_states": [["VA", 99], ["CA", 98], ["OH", 63], ["IL", 50],
                       ["NY", 50], ["PA", 49], ["FL", 42], ["TX", 38],
                       ["NC", 36], ["MI", 24]],
    },
}


def ym(dt):
    """datetime -> 'YYYY-MM' string."""
    return f"{dt.year:04d}-{dt.month:02d}"


def short_name(product_name):
    """Compact chart label from the full Chewy product name."""
    n = product_name
    brand = "FF " if n.startswith("Feline Fresh") else ""
    formula = ""
    for key in ["Healthy", "Multi-Cat", "Pine Pellet", "Unscented",
                "Sisal", "Wide Slatted", "Poop Bags"]:
        if key in n:
            formula = key
            break
    size = ""
    for tok in n.replace(",", " ").split():
        if tok.endswith("-lb") or tok.endswith("lb"):
            size = tok
            break
    if "Poop Bags" in n:
        return "Poop Bags"
    if "Sisal" in n:
        return "Sisal Mat"
    if "Wide Slatted" in n or "Scoop" in n:
        return "Litter Scoop"
    if formula == "Pine Pellet":
        formula = "Pine"
    label = f"{brand}{formula}"
    if size:
        label += f" {size}"
    return label.strip()


def load_flat(wb):
    """Return time series keyed by part number, plus product metadata."""
    ws = wb["FY25_FY26 YTD Data"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(hdr)}

    months = set()
    products = {}
    # A (part, month) can appear in >1 row (Chewy splits published-True/False
    # during a status change). Accumulate: sum units/$, unit-weight the rates.
    acc = defaultdict(lambda: {"units": 0, "net": 0.0,
                               "auto_w": 0.0, "auto_u": 0,
                               "oos_w": 0.0, "oos_u": 0})
    for r in rows[1:]:
        month = r[idx["MONTH"]]
        if month is None:
            continue
        m = ym(month)
        months.add(m)
        part = str(r[idx["PRODUCT_PART_NUMBER"]])
        brand = "Feline Fresh" if str(r[idx["BRAND"]]).startswith("Feline") \
            else "Catalyst"
        if part not in products:
            products[part] = {"name": r[idx["PRODUCT_NAME"]], "brand": brand,
                              "short": short_name(r[idx["PRODUCT_NAME"]]),
                              "ct": None}
        units = int(r[idx["Units Sold"]] or 0)
        a = acc[(part, m)]
        a["units"] += units
        a["net"] += float(r[idx["Net Sales"]] or 0)
        auto = r[idx["Autoship % Units Sold"]]
        if auto is not None:
            a["auto_w"] += float(auto) * units
            a["auto_u"] += units
        oos = r[idx["PDP OOS%"]]
        if oos is not None:
            a["oos_w"] += float(oos) * units
            a["oos_u"] += units

    series = defaultdict(lambda: {"retail": {}, "units": {},
                                  "autoship": {}, "oos": {}})
    for (part, m), a in acc.items():
        series[part]["retail"][m] = round(a["net"], 2)
        series[part]["units"][m] = a["units"]
        if a["auto_u"]:
            series[part]["autoship"][m] = round(a["auto_w"] / a["auto_u"] * 100, 1)
        if a["oos_u"]:
            series[part]["oos"][m] = round(a["oos_w"] / a["oos_u"] * 100, 1)
    return sorted(months), products, series


def load_summary(wb, products, series, months):
    """Attach CT codes + Chewy's FY-YTD retail rollup (authoritative)."""
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    # Monthly column headers live in row index 2, starting at col 12.
    month_cols = {}
    for ci, cell in enumerate(rows[2]):
        if hasattr(cell, "year"):
            month_cols[ym(cell)] = ci

    fy_summary = {}
    used = set()
    for r in rows[3:]:
        ct = r[1]
        desc = r[3]
        if not ct or desc in (None, "") or str(desc).startswith("Total"):
            continue
        # Summary row's monthly retail vector.
        svec = {}
        for m, ci in month_cols.items():
            val = r[ci]
            if val is not None and not isinstance(val, str):
                svec[m] = round(float(val), 2)
        # Best flat part = minimal total abs diff over shared months (unused).
        best, best_score = None, None
        for part, s in series.items():
            if part in used:
                continue
            shared = [m for m in svec if m in s["retail"]]
            if not shared:
                continue
            score = sum(abs(svec[m] - s["retail"][m]) for m in shared) / len(shared)
            if best_score is None or score < best_score:
                best, best_score = part, score
        if best is None:
            continue
        used.add(best)
        products[best]["ct"] = ct
        fy_summary[best] = {
            "fy25": _num(r[4]), "fy26_ytd": _num(r[6]),
            "fy25_ytd": _num(r[7]), "chg_d": _num(r[9]),
            "chg_p": _num(r[10]) if not isinstance(r[10], str) else None,
        }
    return fy_summary


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    months, products, series = load_flat(wb)
    fy_summary = load_summary(wb, products, series, months)

    # ---- Splice in May & June 2026 from the Brand Snapshot PDFs ----
    # Calibrate a realized $/unit per SKU from the trailing Excel actuals
    # (done before overwriting the partial May).
    realized = {}
    for part in products:
        u = sum(series[part]["units"].get(m, 0) for m in PRICE_REF_MONTHS)
        n = sum(series[part]["retail"].get(m, 0) for m in PRICE_REF_MONTHS)
        realized[part] = (n / u) if u else None
    for nm, units_by_part in PDF_MONTHLY_UNITS.items():
        if nm not in months:
            months.append(nm)
        for part, u in units_by_part.items():
            series[part]["units"][nm] = u          # overwrite partial / add
            # Realized $/unit is calibrated from post-changeover (Oct'25+)
            # actuals, so it already reflects current pricing — no extra step.
            rp = realized.get(part)
            if rp:
                series[part]["retail"][nm] = round(rp * u, 2)
            # per-SKU autoship/OOS aren't in the PDFs -> drop any partial value
            series[part]["autoship"].pop(nm, None)
            series[part]["oos"].pop(nm, None)
    months = sorted(months)
    pdf_months = sorted(PDF_MONTHLY_UNITS.keys())

    # Detect a partial trailing month (report pulled mid-month): its total
    # units fall far below the recent run-rate. Excluded from charts/KPIs.
    partial_month = None
    if len(months) >= 4:
        last = months[-1]
        lu = sum(series[p]["units"].get(last, 0) for p in products)
        prior = sorted(sum(series[p]["units"].get(m, 0) for p in products)
                       for m in months[-4:-1])
        med = prior[len(prior) // 2]
        if med and lu < 0.5 * med:
            partial_month = last

    # Compute wholesale $ series (price schedule * units) where priced.
    wholesale = {}
    for part, s in series.items():
        if not WHOLESALE_PRICES.get(part):
            continue
        wholesale[part] = {m: round(wholesale_price(part, m) * u, 2)
                           for m, u in s["units"].items()}
    have_wholesale = bool(wholesale)

    # Implied retail $ = units * shelf (list) price, shown vs actual Net Sales.
    implied = {}
    for part, s in series.items():
        if not SHELF_PRICES.get(part):
            continue
        implied[part] = {m: round(shelf_price(part, m) * u, 2)
                         for m, u in s["units"].items()}

    # Chewy Canada FF panel: units + wholesale $ per available month.
    ca_months = sorted({m for sku in CANADA_FF.values() for m in sku})
    canada = {
        "months": ca_months,
        "units": CANADA_FF,
        "wholesale": {sku: {m: round(CANADA_WHOLESALE[sku] * u, 2)
                            for m, u in months_u.items()}
                      for sku, months_u in CANADA_FF.items()},
    }

    payload = {
        "months": months,
        "products": products,
        "series": series,
        "wholesale": wholesale,
        "have_wholesale": have_wholesale,
        "implied": implied,
        "canada": canada,
        "fy_summary": fy_summary,
        "snapshot": SNAPSHOT,
        "partial_month": partial_month,
        "pdf_months": pdf_months,
        "generated": months[-1],
    }

    with open(TEMPLATE, encoding="utf-8") as f:
        tpl = f.read()
    html = tpl.replace("/*DATA_PLACEHOLDER*/",
                       json.dumps(payload, separators=(",", ":")))
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)

    cat = sum(1 for p in products.values() if p["brand"] == "Catalyst")
    ff = sum(1 for p in products.values() if p["brand"] == "Feline Fresh")
    print(f"[ok] {len(months)} months ({months[0]}..{months[-1]}), "
          f"{len(products)} SKUs ({cat} Catalyst / {ff} Feline Fresh)")
    print(f"[ok] CT codes mapped: "
          f"{sum(1 for p in products.values() if p['ct'])}/{len(products)}")
    print(f"[ok] wholesale prices set: "
          f"{sum(1 for v in WHOLESALE_PRICES.values() if v)}/{len(WHOLESALE_PRICES)}")
    print(f"[ok] PDF months spliced in: {', '.join(pdf_months)}")
    if partial_month:
        print(f"[ok] partial trailing month excluded from charts: {partial_month}")
    print(f"[ok] wrote {OUTPUT}")


if __name__ == "__main__":
    main()
