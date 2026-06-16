"""
Traited / valid status parser.

Reads the latest `Catalyst Store Inventory OOS Summary <MMDDYYYY>.xlsx` snapshot
(Walmart store-inventory pull from Hailey) and produces a per-(store, SKU)
authorization lookup for the dashboard.

"Traited & valid" = the store is on the SKU's official assortment/planogram and
is an active distribution point — i.e. it is *supposed* to carry the item. A
store that appears in the weekly Sales-by-Store feed but is NOT traited has
effectively lost the listing (or never properly set it); its zero-on-hand rows
are an assortment loss, not an actionable out-of-stock. Separating the two is
the whole point of this module.

Snapshot caveat: this is a single point-in-time pull (the date in the filename),
not a weekly time series. Trait status is applied to the dashboard as the
*current* assortment; historical weeks inherit today's status as an approximation.

Output payload (None if no snapshot file is present):
{
  "snapshot_date": "2026-06-15",
  "source_file":   "Catalyst Store Inventory OOS Summary 06152026.xlsx",
  "by_sku":        { "<SKU>": ["<store_num>", ...] },   # traited & valid stores
  "summary":       { "<SKU>"|"Total": {
                        "feed": int,          # rows in the snapshot for the SKU
                        "traited": int,       # traited & valid
                        "non_traited": int,   # in feed, not traited&valid
                        "oos_traited": int,   # traited & on_hand == 0  (true OOS)
                        "oos_non_traited": int# non-traited & on_hand == 0
                     } }
}
"""
from __future__ import annotations

import glob
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent

# Column headers in the snapshot's "Sheet1"
COL_STORE   = "store_number"
COL_DESC    = "all_links_item_description"
COL_TRAITED = "traited_store_count_this_year"
COL_VALID   = "valid_store_count_this_year"
COL_ONHAND  = "store_on_hand_quantity_this_year"

# Our internal SKU keys are the uppercased item descriptions — they match the
# snapshot's description column verbatim (CATALYST15ORIG, CATALYST34LBORIGINAL,
# CATALYST15UNSCEN, CATALYSTPET34LBUNSCE), so no remapping is needed.


def _latest_snapshot_file() -> Path | None:
    candidates = sorted(
        (p for p in glob.glob(str(ROOT / "Catalyst Store Inventory OOS Summary*.xlsx"))
         if not Path(p).name.startswith("~$")),          # skip Excel lock files
        key=lambda p: Path(p).stat().st_mtime,
        reverse=True,
    )
    return Path(candidates[0]) if candidates else None


def _snapshot_date(name: str) -> str | None:
    """Pull MMDDYYYY out of the filename -> ISO date string."""
    m = re.search(r"(\d{2})(\d{2})(\d{4})", name)
    if not m:
        return None
    mm, dd, yyyy = m.groups()
    return f"{yyyy}-{mm}-{dd}"


def build_traited_data(verbose: bool = True) -> dict | None:
    src = _latest_snapshot_file()
    if not src:
        if verbose:
            print("  [traited] No 'Catalyst Store Inventory OOS Summary*.xlsx' found - skipping")
        return None

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header)}
    for need in (COL_STORE, COL_DESC, COL_TRAITED, COL_VALID, COL_ONHAND):
        if need not in idx:
            if verbose:
                print(f"  [traited] '{src.name}' missing column '{need}' - skipping")
            return None

    by_sku: dict[str, list[str]] = {}
    summary: dict[str, dict[str, int]] = {}

    def bump(sku: str, key: str, n: int = 1):
        summary.setdefault(sku, {"feed": 0, "traited": 0, "non_traited": 0,
                                 "oos_traited": 0, "oos_non_traited": 0})
        summary[sku][key] += n

    for r in rows:
        store_raw = r[idx[COL_STORE]]
        desc = r[idx[COL_DESC]]
        if store_raw is None or desc is None:
            continue
        try:
            store = str(int(store_raw))
        except (TypeError, ValueError):
            store = str(store_raw).strip()
        sku = str(desc).strip().upper()

        traited = (r[idx[COL_TRAITED]] == 1) and (r[idx[COL_VALID]] == 1)
        try:
            on_hand = float(r[idx[COL_ONHAND]] or 0)
        except (TypeError, ValueError):
            on_hand = 0.0
        oos = on_hand == 0

        bump(sku, "feed")
        if traited:
            by_sku.setdefault(sku, []).append(store)
            bump(sku, "traited")
            if oos:
                bump(sku, "oos_traited")
        else:
            bump(sku, "non_traited")
            if oos:
                bump(sku, "oos_non_traited")

    # Total rollup
    total = {"feed": 0, "traited": 0, "non_traited": 0, "oos_traited": 0, "oos_non_traited": 0}
    for s in summary.values():
        for k in total:
            total[k] += s[k]
    summary["Total"] = total

    payload = {
        "snapshot_date": _snapshot_date(src.name),
        "source_file":   src.name,
        "by_sku":        by_sku,
        "summary":       summary,
    }

    if verbose:
        t = total
        feed_oos = t["oos_traited"] + t["oos_non_traited"]
        print(f"  [traited] {src.name} (snapshot {payload['snapshot_date']}): "
              f"{t['traited']:,} traited / {t['non_traited']:,} non-traited of {t['feed']:,} feed rows")
        print(f"  [traited]   true OOS (traited & 0 on-hand): {t['oos_traited']} "
              f"({t['oos_traited']/t['traited']*100:.1f}% of traited)  |  "
              f"feed OOS (all): {feed_oos} ({feed_oos/t['feed']*100:.1f}% of feed)")

    return payload


if __name__ == "__main__":
    import json
    data = build_traited_data(verbose=True)
    if data:
        out = ROOT / "traited_debug.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"Wrote {out.name}")
