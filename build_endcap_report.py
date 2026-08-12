# -*- coding: utf-8 -*-
"""
Endcap Rollout Report
=====================
Single self-contained HTML report for the wk202627 endcap set (W/E 08/08/26):
KPIs, status map, week-over-week lift, per-store growth histogram,
prior-distribution segment deep dive, the "still not set" by-reason workbench
(multi-reason filter + Excel export with evidence), and the zero-sale
"set but not selling" flag list.

Inputs:
  - endcap_stores.csv                          (roster, 1,884 stores + flags)
  - (Walmart) Lignetics Inc. Cat Litter Endcap Set WK27.xlsx
        field-visit survey, "Store Details" — the authoritative set/not-set
        status per store, with reason + sub-reason + refusing associate
  - 202626 / 202627 Weekly Sales Report Catalyst.xlsx  (Sales by Store)
  - stores_geo.json                            (zip -> lat/lon cache)

Output: endcap_report.html            (password-gated like dashboard.html)
        endcap-report-x3f8a1.html     (ungated unlisted share copy)

Definitions baked into the page:
  "set" = merchandiser visited and answered Yes to "is the feature product
  set on a feature space" (465 stores). Everything else is either a filed
  reason for not setting (1,394) or a store not yet visited (25). The older
  Exceptions_*.xlsx export is a stale subset of the same survey (1,324 rows)
  and is no longer read.
"""
import csv
import json
import os
from collections import Counter, defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(HERE, "endcap_report.html")
# Ungated share copy: no password, blocked from search engines via
# <meta name="robots" noindex> + an unguessable filename (unlisted URL).
STANDALONE_OUTPUT = os.path.join(HERE, "endcap-report-x3f8a1.html")
PRIOR_WEEK, ENDCAP_WEEK = "202626", "202627"
SURVEY = "(Walmart) Lignetics Inc. Cat Litter Endcap Set WK27.xlsx"

REASONS = {  # survey answer -> (key, label, color)
    "No Available space":                    ("space",     "No available space",   "#d73027"),
    "Not enough inventory to build feature": ("inventory", "Not enough inventory", "#f4a340"),
    "Product not located":                   ("located",   "Product not located",  "#8e44ad"),
    "Store Refusal":                         ("refusal",   "Store refusal",        "#252525"),
}
SET_COLOR, OTHER_COLOR = "#1a9850", "#b8c4cc"
UNVISITED_COLOR = "#7f8c8d"


ENDCAP_UNITS = 36


def load_bystore(week):
    """per-store 15O qty / on-hand / pipeline + all-SKU qty + zip/city/state."""
    wb = openpyxl.load_workbook(
        os.path.join(HERE, f"{week} Weekly Sales Report Catalyst.xlsx"),
        read_only=True)
    q15, oh15, pipe15, qtot, meta = {}, {}, {}, defaultdict(float), {}
    for r in wb["Sales by Store"].iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[1] is None:
            continue
        try:
            store = int(r[1])
        except (TypeError, ValueError):
            continue
        qtot[store] += float(r[8] or 0)
        meta[store] = (str(r[5] or "")[:5], (r[4] or "").title(), r[3] or "")
        if str(r[0]).strip() == "CATALYST15ORIG":
            q15[store] = float(r[8] or 0)
            oh15[store] = float(r[9] or 0)
            pipe15[store] = (float(r[10] or 0), float(r[11] or 0))  # transit, on order
    return q15, oh15, pipe15, qtot, meta


def clean(v):
    """Survey free text arrives partly double-encoded (UTF-8 read as cp1252),
    so 'can’t' lands as 'canâ€™t'. Repair it, leave clean text alone."""
    s = str(v or "").strip()
    if "â€" in s or "Ã" in s:
        try:
            return s.encode("cp1252").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return s
    return s


def load_survey():
    """store -> field-visit record from the WK27 merchandiser survey.

    Columns: 0 store, 1 date, 5 set?, 6 where set, 7 alt-location explain,
    8 priced?, 9 reason not set, 10 not-located sub-reason, 11 not-located
    other, 12 refusal sub-reason, 13 refusal other, 14 refusing associate.
    """
    wb = openpyxl.load_workbook(os.path.join(HERE, SURVEY), read_only=True)
    out = {}
    for r in wb["Store Details"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        try:
            store = int(r[0])
        except (TypeError, ValueError):
            continue
        answer = (r[9] or "").strip()
        key = REASONS.get(answer, (None,))[0]
        # sub-reason: the "other (explain)" free text wins when present
        if key == "located":
            sub, other = clean(r[10]), clean(r[11])
        elif key == "refusal":
            sub, other = clean(r[12]), clean(r[13])
        else:
            sub, other = "", ""
        detail = f"{sub} — {other}".strip(" —") if other else sub
        out[store] = {
            "date": r[1].strftime("%Y-%m-%d") if hasattr(r[1], "strftime") else str(r[1] or ""),
            "set": r[5] == "Yes",
            "where": clean(r[6]),
            "where_other": clean(r[7]),
            "priced": clean(r[8]),
            "seg": "set" if r[5] == "Yes" else (key or "space"),
            "reason": answer,
            "sub": sub,          # dropdown answer only — used for the summary chips
            "detail": detail,    # dropdown answer + free text — used per store
            "title": clean(r[14]),
        }
    return out


def main():
    endcap = {}
    with open(os.path.join(HERE, "endcap_stores.csv"), encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            endcap[int(r["store_number"])] = r

    # --- status: merchandiser survey is authoritative -----------------------
    survey = load_survey()
    seg_of = {}
    for s in endcap:
        v = survey.get(s)
        seg_of[s] = v["seg"] if v else "unvisited"
    reason_label = {key: label for _a, (key, label, _c) in REASONS.items()}

    q15_26, _oh26, _pipe26, qtot_26, meta26 = load_bystore(PRIOR_WEEK)
    q15_27, oh15_27, pipe27, qtot_27, meta27 = load_bystore(ENDCAP_WEEK)

    # endcap product status: has the 36-unit allocation reached the store?
    # "received" counts on-hand + this week's and last week's sell-through
    # (product began arriving in wk202626). No DC-warehouse column in this
    # feed, so DC stock not yet on order shows as "short".
    def inv_status(s):
        if s not in q15_27:
            return "short"
        have = oh15_27.get(s, 0) + q15_27.get(s, 0) + q15_26.get(s, 0)
        transit, onorder = pipe27.get(s, (0, 0))
        if have >= ENDCAP_UNITS:
            return "received"
        if have + transit + onorder >= ENDCAP_UNITS:
            return "transit" if transit > 0 else "onorder"
        return "short"

    set_ok = {s for s in endcap if seg_of[s] == "set"}
    unvisited = {s for s in endcap if seg_of[s] == "unvisited"}
    exc = {s for s in endcap if seg_of[s] not in ("set", "unvisited")}
    visited = set_ok | exc

    # --- lift table --------------------------------------------------------
    def lift_row(label, stores, kind="15O"):
        n = len(stores)
        src26 = q15_26 if kind == "15O" else qtot_26
        src27 = q15_27 if kind == "15O" else qtot_27
        u26 = sum(src26.get(s, 0) for s in stores)
        u27 = sum(src27.get(s, 0) for s in stores)
        return {"label": label, "n": n, "u26": round(u26), "u27": round(u27),
                "usw26": round(u26 / n, 2) if n else 0,
                "usw27": round(u27 / n, 2) if n else 0,
                "pct": round((u27 - u26) / u26 * 100, 1) if u26 else None}

    control = [s for s in q15_26 if s not in endcap]
    lift = [lift_row("Endcap confirmed set", set_ok),
            lift_row("Not set — all reasons", exc)]
    for _answer, (key, label, _c) in REASONS.items():
        lift.append(lift_row("· " + label, {s for s in exc if seg_of[s] == key}))
    lift.append(lift_row("Not visited yet", unvisited))
    lift.append(lift_row("Control: all non-endcap stores", control))
    lift_all = [lift_row("Endcap set — all SKUs", set_ok, "all"),
                lift_row("Not set — all SKUs", exc, "all"),
                lift_row("Control — all SKUs", control, "all")]

    # --- histogram ---------------------------------------------------------
    HBINS = [("No sales either wk", "#c3c2b7"), ("Sold, went to 0", "#a02c2c"),
             ("-99% to -51%", "#d03b3b"), ("-50% to -1%", "#e89392"),
             ("0% (flat)", "#c3c2b7"), ("+1% to +50%", "#9ec5f4"),
             ("+51% to +100%", "#6da7ec"), ("+101% to +200%", "#3987e5"),
             ("over +200%", "#1c5cab"), ("NEW: 0 to selling", "#104281")]
    hist = [0] * 10
    for s in set_ok:
        a, b = q15_26.get(s, 0), q15_27.get(s, 0)
        if a == 0 and b == 0:
            i = 0
        elif a == 0:
            i = 9
        elif b == 0:
            i = 1
        else:
            p = (b - a) / a * 100
            i = (2 if p <= -51 else 3 if p < 0 else 4 if p == 0 else
                 5 if p <= 50 else 6 if p <= 100 else 7 if p <= 200 else 8)
        hist[i] += 1

    # noise baseline from confirmed not-set stores
    exc_up = sum(1 for s in exc if q15_27.get(s, 0) > q15_26.get(s, 0))
    exc_zero27 = sum(1 for s in exc if q15_27.get(s, 0) == 0)
    set_zero27 = sum(1 for s in set_ok if q15_27.get(s, 0) == 0)

    # --- prior-distribution segments --------------------------------------
    no_cat = {s for s, r in endcap.items() if r["on_endcap_no_catalyst"] == "Y"}
    cat_no15 = {s for s, r in endcap.items()
                if r["on_endcap_no_catalyst"] != "Y" and r["has_15O"] != "Y"}
    had15 = {s for s, r in endcap.items() if r["has_15O"] == "Y"}

    def seg_row(label, stores):
        stores = {s for s in stores if s in visited}  # rejection rate over visited only
        n = len(stores)
        rej = [s for s in stores if s in exc]
        ok = [s for s in stores if s in set_ok]
        reasons = Counter(seg_of[s] for s in rej)
        return {"label": label, "n": n, "rej": len(rej),
                "rej_pct": round(len(rej) / n * 100, 1) if n else 0,
                "refusal_pct": round(reasons.get("refusal", 0) / n * 100, 1) if n else 0,
                "located_pct": round(reasons.get("located", 0) / n * 100, 1) if n else 0,
                "n_set": len(ok),
                "set_usw26": round(sum(q15_26.get(s, 0) for s in ok) / len(ok), 2) if ok else 0,
                "set_usw27": round(sum(q15_27.get(s, 0) for s in ok) / len(ok), 2) if ok else 0,
                "set_sold": sum(1 for s in ok if q15_27.get(s, 0) > 0)}
    segments = [seg_row("Already carried the 15-lb", had15),
                seg_row("Carried Catalyst, not the 15-lb", cat_no15),
                seg_row("No Catalyst at all before", no_cat)]
    n333_in27 = sum(1 for s in no_cat if s in q15_27)

    # --- inventory status cross-tab ----------------------------------------
    inv_of = {s: inv_status(s) for s in endcap}
    INV_COLS = ["received", "transit", "onorder", "short"]
    inv_rows = []
    seg_order = [("set", "Confirmed set")] + \
        [(key, label) for _a, (key, label, _c) in REASONS.items()] + \
        [("unvisited", "Not visited yet")]
    for key, label in seg_order:
        stores = [s for s in endcap if seg_of[s] == key]
        row = Counter(inv_of[s] for s in stores)
        inv_rows.append({"label": label, "n": len(stores),
                         **{c: row.get(c, 0) for c in INV_COLS}})
    inv_tot = Counter(inv_of.values())
    rec_stores = [s for s in endcap if inv_of[s] == "received"]
    inv_summary = {
        "counts": {c: inv_tot.get(c, 0) for c in INV_COLS},
        "rec_sold": sum(1 for s in rec_stores if q15_27.get(s, 0) > 0),
        "prev_received": 666,  # from the wk202626 receipts workbook
        "set_noprod": sum(1 for s in set_ok if inv_of[s] != "received"),
        "nei_received": next(r["received"] for r in inv_rows
                             if r["label"] == "Not enough inventory"),
        "nei_n": next(r["n"] for r in inv_rows
                      if r["label"] == "Not enough inventory"),
        "nei_coming": next(r["transit"] + r["onorder"] for r in inv_rows
                           if r["label"] == "Not enough inventory"),
    }

    # --- flag list (zero sales in endcap week) -----------------------------
    flags = []
    for s in sorted(set_ok):
        if q15_27.get(s, 0) > 0:
            continue
        r = endcap[s]
        flags.append({
            "store": s, "city": r["city"].title(), "state": r["state"],
            "tier": 1 if q15_26.get(s, 0) == 0 else 2,
            "q26": round(q15_26.get(s, 0)), "q27": 0,
            "oh": round(oh15_27.get(s, 0)), "inv": inv_of[s],
        })
    flags.sort(key=lambda x: (x["tier"], x["inv"] != "received", -x["oh"]))

    # --- "still not set" workbench: one evidence row per store -------------
    INV_LABEL = {"received": "36 at store (or sold through)", "transit": "In transit",
                 "onorder": "On order only", "short": "Pipeline short of 36"}
    NS_COLS = [
        "Reason", "Reason detail", "Refusing associate", "Visit date",
        "Store #", "Store name", "City", "State", "ZIP", "Address", "Store type",
        "Endcap product status", "On hand 15O", "In transit", "On order",
        f"Units wk{PRIOR_WEEK}", f"Units wk{ENDCAP_WEEK}", "WoW units",
        "Catalyst SKUs before", "Carried 15-lb before", "New to Catalyst",
        "Recommended qty", "Latitude", "Longitude",
    ]
    ns_rows = []
    for s in sorted(exc | unvisited):
        r, v = endcap[s], survey.get(s)
        key = seg_of[s]
        transit, onorder = pipe27.get(s, (0, 0))
        ns_rows.append([
            reason_label.get(key, "Not visited yet"),
            (v or {}).get("detail", ""),
            (v or {}).get("title", ""),
            (v or {}).get("date", ""),
            s, r["store_name"], r["city"].title(), r["state"], str(r["zip"])[:5],
            r["street_address"], r["store_type"],
            INV_LABEL[inv_of[s]], round(oh15_27.get(s, 0)), round(transit), round(onorder),
            round(q15_26.get(s, 0)), round(q15_27.get(s, 0)),
            round(q15_27.get(s, 0) - q15_26.get(s, 0)),
            int(r["current_sku_count"] or 0),
            "Y" if r["has_15O"] == "Y" else "N",
            "Y" if r["on_endcap_no_catalyst"] == "Y" else "N",
            int(r["endcap_recommended_qty"] or 0),
            float(r["latitude"]) if r["latitude"] else "",
            float(r["longitude"]) if r["longitude"] else "",
        ])

    ns_reasons = []
    for key in [k for _a, (k, _l, _c) in REASONS.items()] + ["unvisited"]:
        stores = [s for s in endcap if seg_of[s] == key]
        subs = Counter(survey[s]["sub"] for s in stores
                       if survey.get(s) and survey[s]["sub"])
        ns_reasons.append({
            "key": key,
            "label": reason_label.get(key, "Not visited yet"),
            "color": next((c for _a, (k, _l, c) in REASONS.items() if k == key),
                          UNVISITED_COLOR),
            "n": len(stores),
            "received": sum(1 for s in stores if inv_of[s] == "received"),
            "inbound": sum(1 for s in stores if inv_of[s] in ("transit", "onorder")),
            "subs": subs.most_common(),
        })

    # --- map stores --------------------------------------------------------
    geo = json.load(open(os.path.join(HERE, "stores_geo.json"), encoding="utf-8"))
    map_stores = []
    for s, r in endcap.items():
        if r["latitude"] and r["longitude"]:
            lat, lon = float(r["latitude"]), float(r["longitude"])
        else:
            g = geo.get(str(r["zip"])[:5]) or geo.get(meta27.get(s, ("",))[0])
            if not g:
                continue
            lat, lon = g["lat"], g["lon"]
        prior = ("none" if r["on_endcap_no_catalyst"] == "Y"
                 else f"{r['current_sku_count']} SKU" +
                      ("" if r["current_sku_count"] == "1" else "s") +
                      (", no 15-lb" if r["has_15O"] != "Y" else ""))
        map_stores.append({"store": s, "city": r["city"].title(), "state": r["state"],
                           "lat": round(lat, 4), "lon": round(lon, 4),
                           "seg": seg_of[s], "prior": prior,
                           "q26": q15_26.get(s, 0), "q27": q15_27.get(s, 0)})
    for s in qtot_27:
        if s in endcap:
            continue
        zip5, city, state = meta27.get(s, ("", "", ""))
        g = geo.get(zip5)
        if not g:
            continue
        map_stores.append({"store": s, "city": city, "state": state,
                           "lat": g["lat"], "lon": g["lon"], "seg": "other",
                           "prior": None,
                           "q26": q15_26.get(s, 0), "q27": q15_27.get(s, 0)})

    seg_meta = {"set": {"label": "Endcap confirmed set", "color": SET_COLOR}}
    for _answer, (key, label, color) in REASONS.items():
        seg_meta[key] = {"label": label, "color": color}
    seg_meta["unvisited"] = {"label": "Not visited yet", "color": UNVISITED_COLOR}
    seg_meta["other"] = {"label": "All other Catalyst stores", "color": OTHER_COLOR}
    seg_counts = {k: sum(1 for x in map_stores if x["seg"] == k) for k in seg_meta}

    inc_units = round(lift[0]["u27"] - lift[0]["u26"])
    payload = {
        "prior_week": PRIOR_WEEK, "week": ENDCAP_WEEK,
        "n_endcap": len(endcap), "n_set": len(set_ok), "n_exc": len(exc),
        "n_unvisited": len(unvisited), "n_visited": len(visited),
        "confirmed_set": len(set_ok),
        "set_where": dict(Counter(survey[s]["where"] for s in set_ok if survey.get(s))),
        "set_priced": sum(1 for s in set_ok if survey.get(s, {}).get("priced") == "Yes"),
        "kpi": {"lift_pct": lift[0]["pct"], "exc_pct": round(len(exc) / len(visited) * 100),
                "inc_units": inc_units, "inc_retail": round(inc_units * 15.97)},
        "notset": {"cols": NS_COLS, "rows": ns_rows, "reasons": ns_reasons},
        "lift": lift, "lift_all": lift_all,
        "hist": {"labels": [l for l, _ in HBINS], "colors": [c for _, c in HBINS],
                 "counts": hist},
        "noise": {"exc_up_pct": round(exc_up / len(exc) * 100),
                  "exc_zero_pct": round(exc_zero27 / len(exc) * 100),
                  "set_zero": set_zero27,
                  "set_zero_pct": round(set_zero27 / len(set_ok) * 100)},
        "segments": segments, "n333": len(no_cat), "n333_in27": n333_in27,
        "inv": {"rows": inv_rows, **inv_summary, "units": ENDCAP_UNITS},
        "flags": flags,
        "map": {"segments": seg_meta, "counts": seg_counts, "stores": map_stores},
    }

    html = TEMPLATE.replace("/*DATA*/", json.dumps(payload, separators=(",", ":")))
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[ok] set={len(set_ok)} notset={len(exc)} unvisited={len(unvisited)} "
          f"flags={len(flags)} notset_rows={len(ns_rows)} map_stores={len(map_stores)}")
    print(f"[ok] wrote {OUTPUT}")

    # ungated share copy: strip the password gate, keep search engines out
    standalone = (html.replace(GATE_HTML, "")
                      .replace(GATE_JS, GATE_JS_OPEN)
                      .replace(GATE_TAIL, "showApp();"))
    for frag in (GATE_HTML, GATE_JS, GATE_TAIL):
        if frag in standalone:
            raise SystemExit("[err] gate strip failed — template drifted")
    with open(STANDALONE_OUTPUT, "w", encoding="utf-8") as f:
        f.write(standalone)
    print(f"[ok] wrote {STANDALONE_OUTPUT} (no gate, noindex)")


GATE_HTML = """<div id="gate"><div class="box">
  <div style="font-size:18px;font-weight:700">Catalyst Endcap Report</div>
  <div style="color:var(--ink2);font-size:13px;margin-top:4px">Enter password to continue</div>
  <input type="password" id="pw" placeholder="Password" autocomplete="current-password">
  <div class="err" id="pwerr"></div>
</div></div>
"""

GATE_JS = """/* ---- password gate (same key as dashboard.html) ---- */
const AUTH_KEY = "wm_catalyst_auth", PASSWORD = "pellets123";
function showApp(){ document.getElementById('gate').style.display='none';
  document.getElementById('app').style.display='block'; initMap(); }
document.getElementById('pw').addEventListener('keydown', e => {
  if (e.key !== 'Enter') return;
  if (e.target.value === PASSWORD) { localStorage.setItem(AUTH_KEY,'ok'); showApp(); }
  else document.getElementById('pwerr').textContent = 'Incorrect password. Try again.';
});"""

GATE_JS_OPEN = ("function showApp(){ document.getElementById('app')"
                ".style.display='block'; initMap(); }")

GATE_TAIL = """if (localStorage.getItem(AUTH_KEY) === 'ok') showApp();
else document.getElementById('pw').focus();"""

TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex, nofollow, noarchive">
<title>Catalyst Endcap Rollout Report</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {
    --surface:#fcfcfb; --page:#f4f4f1; --ink:#0b0b0b; --ink2:#52514e;
    --muted:#898781; --grid:#e1e0d9; --line:#c3c2b7; --green:#1a9850;
  }
  * { box-sizing:border-box; }
  html, body { margin:0; background:var(--page); color:var(--ink);
    font-family:system-ui,-apple-system,"Segoe UI",sans-serif; }
  .wrap { max-width:1100px; margin:0 auto; padding:24px 20px 60px; }
  header h1 { font-size:24px; margin:0 0 4px; }
  header .sub { color:var(--ink2); font-size:14px; }
  section { background:var(--surface); border:1px solid var(--grid);
    border-radius:10px; padding:20px 24px; margin-top:20px; }
  section h2 { font-size:17px; margin:0 0 4px; }
  section .note { color:var(--muted); font-size:12.5px; margin:0 0 14px; }
  .kpis { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr));
    gap:12px; margin-top:20px; }
  .kpi { background:var(--surface); border:1px solid var(--grid); border-radius:10px;
    padding:14px 16px; }
  .kpi .v { font-size:26px; font-weight:700; }
  .kpi .l { color:var(--ink2); font-size:12.5px; margin-top:2px; }
  .kpi .d { color:var(--muted); font-size:11.5px; margin-top:2px; }
  #map { height:520px; border-radius:8px; }
  .maplegend { display:flex; flex-wrap:wrap; gap:6px 18px; margin-top:10px; font-size:13px; }
  .maplegend .row { display:flex; align-items:center; gap:6px; cursor:pointer;
    user-select:none; }
  .maplegend .row.off { opacity:.35; }
  .maplegend .dot { width:11px; height:11px; border-radius:50%; }
  .maplegend .n { color:var(--muted); }
  table { border-collapse:collapse; width:100%; font-size:13.5px; }
  th { text-align:left; color:var(--ink2); font-weight:600; font-size:12px;
    text-transform:uppercase; letter-spacing:.03em; border-bottom:2px solid var(--line);
    padding:6px 10px; }
  td { padding:6px 10px; border-bottom:1px solid var(--grid); }
  td.num, th.num { text-align:right; font-variant-numeric:tabular-nums; }
  tr.hl td { background:#eef7ee; font-weight:600; }
  tr.dim td { color:var(--ink2); }
  .pos { color:#006300; font-weight:600; }
  .neg { color:#b3261e; font-weight:600; }
  .chart { width:100%; }
  .bars { display:flex; align-items:flex-end; gap:10px; height:250px; margin-top:26px; }
  .bar { flex:1; display:flex; flex-direction:column; align-items:center;
    justify-content:flex-end; height:100%; }
  .bar .fill { width:100%; max-width:74px; border-radius:4px 4px 0 0; }
  .bar .cnt { font-weight:700; font-size:14px; margin-bottom:2px; }
  .bar .pct { color:var(--muted); font-size:11px; }
  .bar .lab { color:var(--ink2); font-size:11px; text-align:center; margin-top:6px;
    line-height:1.25; height:30px; }
  .callout { background:#f6f4ef; border-left:3px solid var(--line); border-radius:6px;
    padding:10px 14px; font-size:13.5px; color:var(--ink2); margin-top:14px; }
  .callout b { color:var(--ink); }
  .flagwrap { max-height:420px; overflow:auto; border:1px solid var(--grid);
    border-radius:8px; margin-top:10px; }
  .flagwrap th { position:sticky; top:0; background:var(--surface); }
  .btn { display:inline-block; background:var(--ink); color:#fff; border:none;
    border-radius:6px; padding:7px 14px; font-size:13px; cursor:pointer; }
  .tier1 { background:#fdecea; border-radius:4px; padding:1px 7px; font-size:12px; }
  .tier2 { background:#fef3e2; border-radius:4px; padding:1px 7px; font-size:12px; }
  /* not-set workbench */
  .picker { display:grid; grid-template-columns:repeat(auto-fit,minmax(250px,1fr));
    gap:10px; margin:14px 0 4px; }
  .pick { display:flex; align-items:center; gap:10px; border:1px solid var(--grid);
    border-radius:8px; padding:10px 12px; cursor:pointer; user-select:none;
    background:var(--surface); }
  .pick:hover { border-color:var(--line); }
  .pick.on { border-color:var(--ink); background:#f6f7f4; }
  .pick input { width:16px; height:16px; accent-color:var(--ink); cursor:pointer; }
  .pick .swatch { width:10px; height:10px; border-radius:50%; flex:none; }
  .pick .txt { flex:1; }
  .pick .txt b { display:block; font-size:13.5px; }
  .pick .txt span { color:var(--muted); font-size:11.5px; }
  .pick .cnt { font-size:17px; font-weight:700; font-variant-numeric:tabular-nums; }
  .toolbar { display:flex; flex-wrap:wrap; align-items:center; gap:10px; margin-top:14px; }
  .toolbar .sel { font-size:13.5px; color:var(--ink2); flex:1; min-width:220px; }
  .toolbar .sel b { color:var(--ink); }
  .btn.ghost { background:transparent; color:var(--ink); border:1px solid var(--line); }
  .btn:disabled { opacity:.4; cursor:not-allowed; }
  .subs { margin-top:12px; font-size:12.5px; color:var(--ink2); }
  .subs .grp { margin-top:6px; }
  .subs .chip { display:inline-block; background:#f1f0ea; border-radius:20px;
    padding:2px 10px; margin:3px 4px 0 0; font-size:12px; }
  .nswrap { max-height:460px; overflow:auto; border:1px solid var(--grid);
    border-radius:8px; margin-top:12px; }
  .nswrap th { position:sticky; top:0; background:var(--surface); z-index:2; }
  .nswrap td { white-space:nowrap; }
  .rdot { display:inline-block; width:8px; height:8px; border-radius:50%;
    margin-right:6px; vertical-align:middle; }
  #nssearch { font-size:13px; padding:6px 10px; border:1px solid var(--line);
    border-radius:6px; width:210px; }
  footer { color:var(--muted); font-size:12px; margin-top:26px; line-height:1.5; }
  /* password gate */
  #gate { position:fixed; inset:0; background:var(--page); z-index:5000;
    display:flex; align-items:center; justify-content:center; }
  #gate .box { background:var(--surface); border:1px solid var(--grid);
    border-radius:12px; padding:34px 38px; text-align:center; }
  #gate input { font-size:15px; padding:8px 12px; border:1px solid var(--line);
    border-radius:6px; margin-top:12px; width:220px; }
  #gate .err { color:#b3261e; font-size:13px; height:18px; margin-top:8px; }
  #app { display:none; }
</style>
</head>
<body>
<div id="gate"><div class="box">
  <div style="font-size:18px;font-weight:700">Catalyst Endcap Report</div>
  <div style="color:var(--ink2);font-size:13px;margin-top:4px">Enter password to continue</div>
  <input type="password" id="pw" placeholder="Password" autocomplete="current-password">
  <div class="err" id="pwerr"></div>
</div></div>

<div id="app"><div class="wrap">
<header>
  <h1>Catalyst 15-lb Endcap Rollout Report</h1>
  <div class="sub" id="subtitle"></div>
</header>

<div class="kpis" id="kpis"></div>

<section>
  <h2>Where the endcap landed — and where it didn&rsquo;t</h2>
  <p class="note">1,884 program stores colored by outcome; grey dots are every other store selling Catalyst this week. Click legend entries to toggle layers; click a pin for the store&rsquo;s week-over-week units.</p>
  <div id="map"></div>
  <div class="maplegend" id="maplegend"></div>
</section>

<section>
  <h2>Week-over-week lift — 15-lb Original units per store (U/S/W)</h2>
  <p class="note">Week 202626 (before the set) vs. week 202627 (endcap week). &ldquo;Set&rdquo; = no exception filed.</p>
  <table id="lifttable"></table>
  <div class="callout" id="liftcallout"></div>
</section>

<section>
  <h2>Growth distribution across the confirmed-set stores</h2>
  <p class="note" id="histnote">Week-over-week % change in 15-lb units, one bar segment per store outcome. Red = declined, grey = no signal, blue = grew.</p>
  <div class="bars" id="hist"></div>
  <div class="callout" id="histcallout"></div>
</section>

<section>
  <h2>Has the endcap product physically arrived? (wk202627 inventory)</h2>
  <p class="note" id="invnote"></p>
  <table id="invtable"></table>
  <div class="callout" id="invcallout"></div>
</section>

<section>
  <h2>Did stores new to Catalyst reject the endcap more?</h2>
  <p class="note">The roster splits by what each store carried before the program. Rejection = any exception filed.</p>
  <table id="segtable"></table>
  <div class="callout" id="segcallout"></div>
</section>

<section>
  <h2>Stores still not set — pick reasons, pull the list</h2>
  <p class="note" id="nsnote"></p>
  <div class="picker" id="nspicker"></div>
  <div class="toolbar">
    <div class="sel" id="nssel"></div>
    <input id="nssearch" type="search" placeholder="Filter: store #, city, state&hellip;">
    <button class="btn ghost" id="nsall">Select all</button>
    <button class="btn ghost" id="nsnone">Clear</button>
    <button class="btn" id="nsxlsx">Download Excel</button>
    <button class="btn ghost" id="nscsv">CSV</button>
  </div>
  <div class="subs" id="nssubs"></div>
  <div class="nswrap"><table id="nstable"></table></div>
  <p class="note" id="nsmore"></p>
</section>

<section>
  <h2>Flag list — confirmed-set stores with zero endcap-week sales</h2>
  <p class="note" id="flagnote"></p>
  <button class="btn" id="csvbtn">Download CSV</button>
  <div class="flagwrap"><table id="flagtable"></table></div>
</section>

<footer id="foot"></footer>
</div></div>

<script>
const DATA = /*DATA*/;

/* ---- password gate (same key as dashboard.html) ---- */
const AUTH_KEY = "wm_catalyst_auth", PASSWORD = "pellets123";
function showApp(){ document.getElementById('gate').style.display='none';
  document.getElementById('app').style.display='block'; initMap(); }
document.getElementById('pw').addEventListener('keydown', e => {
  if (e.key !== 'Enter') return;
  if (e.target.value === PASSWORD) { localStorage.setItem(AUTH_KEY,'ok'); showApp(); }
  else document.getElementById('pwerr').textContent = 'Incorrect password. Try again.';
});

const fmt = n => n.toLocaleString();
const wkLabel = w => 'wk ' + w;

document.getElementById('subtitle').textContent =
  DATA.n_endcap.toLocaleString() + ' program stores · set W/E 08/08/26 · sales ' +
  wkLabel(DATA.prior_week) + ' → ' + wkLabel(DATA.week) + ' · generated from Walmart weekly data';

/* ---- KPIs ---- */
const K = DATA.kpi;
/* NB: never name anything `L` here — Leaflet owns that global. */
const LIFT = DATA.lift, byLabel = l => LIFT.find(r => r.label === l);
const notSetPct = p => (p === null ? '—' : (p>0?'+':'') + p.toFixed(1) + '%');
const kpis = [
  [fmt(DATA.n_endcap), 'endcap program stores',
   fmt(DATA.n_visited) + ' visited · ' + fmt(DATA.n_unvisited) + ' not yet'],
  [fmt(DATA.n_set) + ' (' + Math.round(100*DATA.n_set/DATA.n_visited) + '%)',
   'confirmed set by a merchandiser', DATA.set_priced + ' of them priced correctly'],
  [fmt(DATA.n_exc) + ' (' + K.exc_pct + '%)', 'visited but not set', 'reason filed for every one'],
  ['+' + K.lift_pct + '%', 'U/S/W lift in set stores',
   'vs ' + notSetPct(LIFT[1].pct) + ' not set, ' +
   notSetPct(byLabel('Control: all non-endcap stores').pct) + ' chain control'],
  ['+' + fmt(K.inc_units) + ' units', 'incremental 15-lb units, week one',
   '≈ $' + fmt(K.inc_retail) + ' retail']
];
document.getElementById('kpis').innerHTML = kpis.map(k =>
  '<div class="kpi"><div class="v">' + k[0] + '</div><div class="l">' + k[1] +
  '</div>' + (k[2] ? '<div class="d">' + k[2] + '</div>' : '') + '</div>').join('');

/* ---- map ---- */
let mapInit = false;
function initMap(){
  if (mapInit) return; mapInit = true;
  const map = L.map('map').setView([38.8, -93], 4);
  L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',
    { attribution:'&copy; OpenStreetMap &copy; CARTO', maxZoom:18 }).addTo(map);
  const groups = {};
  for (const key in DATA.map.segments) groups[key] = L.layerGroup().addTo(map);
  for (const s of DATA.map.stores) {
    const seg = DATA.map.segments[s.seg], bg = s.seg === 'other';
    const m = L.circleMarker([s.lat, s.lon], { radius: bg?3:5, color:'#333',
      weight: bg?0:.5, fillColor: seg.color, fillOpacity: bg?.4:.85 });
    let rows = '<tr><td>15-lb units ' + wkLabel(DATA.prior_week) + '</td><td>' + s.q26 + '</td></tr>' +
               '<tr><td>15-lb units ' + wkLabel(DATA.week) + '</td><td><b>' + s.q27 + '</b></td></tr>';
    if (s.prior !== null) rows += '<tr><td>Catalyst before endcap</td><td>' + s.prior + '</td></tr>';
    m.bindPopup('<b>Store #' + s.store + '</b> — ' + s.city + ', ' + s.state +
      '<br><span style="color:' + seg.color + ';font-weight:600">' + seg.label + '</span>' +
      '<table style="font-size:12px">' + rows + '</table>');
    m.addTo(groups[s.seg]);
  }
  const leg = document.getElementById('maplegend');
  for (const key in DATA.map.segments) {
    const seg = DATA.map.segments[key], n = DATA.map.counts[key];
    const pct = key === 'other' ? '' : ' (' + Math.round(100*n/DATA.n_endcap) + '%)';
    const row = document.createElement('div');
    row.className = 'row';
    row.innerHTML = '<span class="dot" style="background:' + seg.color + '"></span>' +
      seg.label + '<span class="n">' + fmt(n) + pct + '</span>';
    row.onclick = () => { const off = row.classList.toggle('off');
      off ? map.removeLayer(groups[key]) : map.addLayer(groups[key]); };
    leg.appendChild(row);
  }
}

/* ---- lift table ---- */
function pctCell(p){
  if (p === null) return '<td class="num">—</td>';
  const cls = p > 5 ? 'pos' : p < -5 ? 'neg' : '';
  return '<td class="num ' + cls + '">' + (p>0?'+':'') + p.toFixed(1) + '%</td>';
}
let lt = '<tr><th>Group</th><th class="num">Stores</th>' +
  '<th class="num">Units ' + wkLabel(DATA.prior_week) + '</th>' +
  '<th class="num">Units ' + wkLabel(DATA.week) + '</th>' +
  '<th class="num">U/S/W before</th><th class="num">U/S/W after</th><th class="num">Change</th></tr>';
DATA.lift.forEach((r, i) => {
  const cls = i === 0 ? 'hl' : r.label.startsWith('·') ? 'dim' : '';
  lt += '<tr class="' + cls + '"><td>' + r.label + '</td><td class="num">' + fmt(r.n) +
    '</td><td class="num">' + fmt(r.u26) + '</td><td class="num">' + fmt(r.u27) +
    '</td><td class="num">' + r.usw26.toFixed(2) + '</td><td class="num">' +
    r.usw27.toFixed(2) + '</td>' + pctCell(r.pct) + '</tr>';
});
document.getElementById('lifttable').innerHTML = lt;
const la = DATA.lift_all;
document.getElementById('liftcallout').innerHTML =
  '<b>The lift is real growth, not shelf cannibalization:</b> all-SKU units in set stores rose ' +
  '+' + la[0].pct + '% (' + fmt(la[0].u26) + ' → ' + fmt(la[0].u27) + ') while not-set stores were +' +
  la[1].pct + '% and the chain control ' + (la[2].pct>0?'+':'') + la[2].pct + '%. ' +
  '&ldquo;Product not located&rdquo; and &ldquo;store refusal&rdquo; stores actually declined.';

/* ---- histogram ---- */
const H = DATA.hist, maxH = Math.max(...H.counts);
document.getElementById('hist').innerHTML = H.counts.map((n, i) =>
  '<div class="bar"><div class="cnt">' + n + '</div><div class="pct">' +
  Math.round(100*n/DATA.n_set) + '%</div><div class="fill" style="height:' +
  Math.max(2, n/maxH*170) + 'px;background:' + H.colors[i] + '"></div>' +
  '<div class="lab">' + H.labels[i] + '</div></div>').join('');
const NZ = DATA.noise;
document.getElementById('histnote').textContent =
  'Week-over-week % change in 15-lb units across the ' + fmt(DATA.n_set) +
  ' stores a merchandiser confirmed set. Red = declined, grey = no signal, blue = grew.';
document.getElementById('histcallout').innerHTML =
  '<b>How to read this:</b> store-level weeks are noisy (typical store sells 1–3 units), ' +
  'so even among confirmed not-set stores ' + NZ.exc_up_pct + '% ticked up. The cleaner tell is ' +
  '<b>zero sales in the endcap week</b>: ' + NZ.set_zero + ' confirmed-set stores (' + NZ.set_zero_pct +
  '%) sold nothing vs ' + NZ.exc_zero_pct + '% of not-set stores. Since these ' + fmt(DATA.n_set) +
  ' were visited and photographed as set, a zero week means the endcap is up but not moving — ' +
  'they are the flag list at the bottom of the page.';

/* ---- inventory status ---- */
const INV = DATA.inv;
const INV_LABELS = {received:'36 at store (or sold through)', transit:'In transit',
  onorder:'On order only', short:'Pipeline short of 36'};
document.getElementById('invnote').textContent =
  'Each program store was allocated ' + INV.units + ' units. “36 at store” counts on-hand plus the last two weeks’ ' +
  'sell-through, so stores that received and sold bags still count. This feed has no DC-warehouse column, so DC ' +
  'stock not yet on a store order shows as “short”.';
let it = '<tr><th>Set status</th><th class="num">Stores</th>' +
  Object.values(INV_LABELS).map(l => '<th class="num">' + l + '</th>').join('') +
  '<th class="num">% arrived</th></tr>';
DATA.inv.rows.forEach((r, i) => {
  it += '<tr class="' + (i===0?'hl':'') + '"><td>' + r.label + '</td><td class="num">' + fmt(r.n) + '</td>' +
    ['received','transit','onorder','short'].map(c => '<td class="num">' + fmt(r[c]) + '</td>').join('') +
    '<td class="num">' + Math.round(100*r.received/r.n) + '%</td></tr>';
});
const IC = INV.counts, invTot = IC.received + IC.transit + IC.onorder + IC.short;
it += '<tr style="font-weight:600"><td>All program stores</td><td class="num">' + fmt(invTot) + '</td>' +
  ['received','transit','onorder','short'].map(c => '<td class="num">' + fmt(IC[c]) + '</td>').join('') +
  '<td class="num">' + Math.round(100*IC.received/invTot) + '%</td></tr>';
document.getElementById('invtable').innerHTML = it;
document.getElementById('invcallout').innerHTML =
  '<b>' + fmt(IC.received) + ' stores (' + Math.round(100*IC.received/invTot) + '%) now have their 36 bags</b> (' +
  fmt(INV.rec_sold) + ' of them sold at least one this week), up from ' + fmt(INV.prev_received) +
  ' in last week&rsquo;s receipts file; ' + fmt(IC.transit) + ' are in transit and ' + fmt(IC.onorder) +
  ' still on order. The reason codes check out against the pipeline: &ldquo;not enough inventory&rdquo; stores are the ' +
  'least stocked — only ' + INV.nei_received + ' of ' + fmt(INV.nei_n) + ' have product even now, with ' +
  fmt(INV.nei_coming) + ' inbound — so most of those can set once trucks land. Read the top row with care: ' +
  fmt(INV.set_noprod) + ' stores a merchandiser photographed as set still show fewer than ' + INV.units +
  ' units received, which is receipt-posting lag or on-hand inaccuracy rather than a missing display.';

/* ---- segments ---- */
let st = '<tr><th>Prior distribution</th><th class="num">Stores</th>' +
  '<th class="num">Rejected</th><th class="num">Rejection rate</th>' +
  '<th class="num">Outright refusal</th><th class="num">Product not located</th>' +
  '<th class="num">Set: U/S/W before → after</th><th class="num">Set &amp; sold &gt;0</th></tr>';
DATA.segments.forEach(r => {
  st += '<tr><td>' + r.label + '</td><td class="num">' + fmt(r.n) + '</td><td class="num">' +
    fmt(r.rej) + '</td><td class="num">' + r.rej_pct + '%</td><td class="num">' +
    r.refusal_pct + '%</td><td class="num">' + r.located_pct + '%</td><td class="num">' +
    r.set_usw26.toFixed(2) + ' → ' + r.set_usw27.toFixed(2) + '</td><td class="num">' +
    r.set_sold + ' / ' + r.n_set + '</td></tr>';
});
document.getElementById('segtable').innerHTML = st;
document.getElementById('segcallout').innerHTML =
  '<b>Yes — and it scales with unfamiliarity.</b> Stores that never carried Catalyst rejected at ' +
  DATA.segments[2].rej_pct + '% vs ' + DATA.segments[0].rej_pct + '% for stores already selling the 15-lb; ' +
  'their outright-refusal rate runs ' + (DATA.segments[2].refusal_pct/DATA.segments[0].refusal_pct).toFixed(1) +
  '× higher and &ldquo;product not located&rdquo; ' +
  (DATA.segments[2].located_pct/DATA.segments[0].located_pct).toFixed(1) + '× higher. ' +
  'All ' + DATA.n333 + ' never-stocked stores DO appear in Hailey&rsquo;s weekly file (' + DATA.n333_in27 +
  ' of ' + DATA.n333 + ' have a 15-lb row in ' + wkLabel(DATA.week) + ' — they were traited when the ' +
  'endcap product shipped), so data coverage is not the issue. Their week-one U/S/W of ' +
  DATA.segments[2].set_usw27.toFixed(2) + ' in set stores is promising for zero-history doors, but well under the ' +
  DATA.segments[0].set_usw27.toFixed(2) + ' of established ones.';

/* ---- not-set workbench ---- */
const NS = DATA.notset, NSC = NS.cols;
const C_REASON = 0, C_DETAIL = 1, C_TITLE = 2, C_DATE = 3, C_STORE = 4,
      C_CITY = 6, C_STATE = 7, C_INV = 11, C_OH = 12, C_U26 = 15, C_U27 = 16;
const nsColor = {}; NS.reasons.forEach(r => nsColor[r.label] = r.color);
const nsSel = new Set(NS.reasons.filter(r => r.key !== 'unvisited').map(r => r.key));
const MAXROWS = 400;

document.getElementById('nsnote').textContent =
  'Every program store that is not confirmed set, with the reason the merchandiser filed on the ' +
  'W/E 08/08/26 visit. Tick one or more reasons to combine them — the table and both downloads ' +
  'follow the selection. The Excel file carries all ' + NSC.length + ' evidence columns per store: ' +
  'reason and sub-reason, refusing associate, visit date, full address, endcap-product status, ' +
  'on-hand and inbound units, and the before/after weekly sales.';

document.getElementById('nspicker').innerHTML = NS.reasons.map(r =>
  '<label class="pick' + (nsSel.has(r.key) ? ' on' : '') + '" data-key="' + r.key + '">' +
  '<input type="checkbox"' + (nsSel.has(r.key) ? ' checked' : '') + '>' +
  '<span class="swatch" style="background:' + r.color + '"></span>' +
  '<span class="txt"><b>' + r.label + '</b><span>' + fmt(r.received) + ' have the product · ' +
  fmt(r.inbound) + ' inbound</span></span>' +
  '<span class="cnt">' + fmt(r.n) + '</span></label>').join('');

function nsRows(){
  const labels = new Set(NS.reasons.filter(r => nsSel.has(r.key)).map(r => r.label));
  const q = document.getElementById('nssearch').value.trim().toLowerCase();
  return NS.rows.filter(r => labels.has(r[C_REASON]) &&
    (!q || (r[C_STORE] + ' ' + r[C_CITY] + ' ' + r[C_STATE] + ' ' + r[C_DETAIL]).toLowerCase().includes(q)));
}
function selLabel(){
  const on = NS.reasons.filter(r => nsSel.has(r.key));
  return on.length === 0 ? 'nothing selected'
       : on.length === NS.reasons.length ? 'every not-set store'
       : on.map(r => r.label.toLowerCase()).join(' + ');
}
function renderNS(){
  const rows = nsRows();
  document.getElementById('nssel').innerHTML =
    '<b>' + fmt(rows.length) + '</b> stores — ' + selLabel();
  document.getElementById('nsxlsx').disabled = !rows.length;
  document.getElementById('nscsv').disabled = !rows.length;

  const subs = NS.reasons.filter(r => nsSel.has(r.key) && r.subs.length);
  document.getElementById('nssubs').innerHTML = subs.map(r =>
    '<div class="grp"><b>' + r.label + '</b> — reasons given: ' +
    r.subs.map(s => '<span class="chip">' + s[0] + ' · ' + fmt(s[1]) + '</span>').join('') +
    '</div>').join('');

  const show = rows.slice(0, MAXROWS);
  let h = '<tr><th>Reason</th><th>Store #</th><th>City</th><th>State</th><th>Detail</th>' +
    '<th>Endcap product</th><th class="num">On hand</th>' +
    '<th class="num">Units ' + wkLabel(DATA.prior_week) + '</th>' +
    '<th class="num">Units ' + wkLabel(DATA.week) + '</th><th>Visited</th></tr>';
  show.forEach(r => {
    h += '<tr><td><span class="rdot" style="background:' + (nsColor[r[C_REASON]] || '#999') +
      '"></span>' + r[C_REASON] + '</td><td>' + r[C_STORE] + '</td><td>' + r[C_CITY] +
      '</td><td>' + r[C_STATE] + '</td><td>' + (r[C_DETAIL] || '—') + '</td><td>' + r[C_INV] +
      '</td><td class="num">' + r[C_OH] + '</td><td class="num">' + r[C_U26] +
      '</td><td class="num">' + r[C_U27] + '</td><td>' + (r[C_DATE] || '—') + '</td></tr>';
  });
  document.getElementById('nstable').innerHTML = h;
  document.getElementById('nsmore').textContent = rows.length > MAXROWS
    ? 'Showing the first ' + MAXROWS + ' of ' + fmt(rows.length) +
      ' — the download contains all ' + fmt(rows.length) + '.' : '';
}
document.getElementById('nspicker').addEventListener('change', e => {
  const lab = e.target.closest('.pick'), key = lab.dataset.key;
  e.target.checked ? nsSel.add(key) : nsSel.delete(key);
  lab.classList.toggle('on', e.target.checked);
  renderNS();
});
document.getElementById('nssearch').addEventListener('input', renderNS);
function setAll(on){
  NS.reasons.forEach(r => on ? nsSel.add(r.key) : nsSel.delete(r.key));
  document.querySelectorAll('#nspicker .pick').forEach(l => {
    l.querySelector('input').checked = on; l.classList.toggle('on', on); });
  renderNS();
}
document.getElementById('nsall').onclick = () => setAll(true);
document.getElementById('nsnone').onclick = () => setAll(false);

function nsFilename(ext){
  const on = NS.reasons.filter(r => nsSel.has(r.key));
  const tag = on.length === NS.reasons.length ? 'all'
    : on.map(r => r.key).join('-') || 'none';
  return 'catalyst_endcap_not_set_' + tag + '_wk' + DATA.week + '.' + ext;
}
function download(blob, name){
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob); a.download = name; a.click();
  setTimeout(() => URL.revokeObjectURL(a.href), 1000);
}
document.getElementById('nscsv').onclick = () => {
  const esc = v => { v = String(v == null ? '' : v);
    return /[",\n]/.test(v) ? '"' + v.replace(/"/g, '""') + '"' : v; };
  const csv = [NSC, ...nsRows()].map(r => r.map(esc).join(',')).join('\n');
  download(new Blob([csv], {type:'text/csv;charset=utf-8'}), nsFilename('csv'));
};
document.getElementById('nsxlsx').onclick = () => {
  if (typeof XLSX === 'undefined') {
    alert('The Excel library did not load (offline?). Use the CSV button instead.');
    return;
  }
  const rows = nsRows();
  const wb = XLSX.utils.book_new();

  const ws = XLSX.utils.aoa_to_sheet([NSC, ...rows]);
  ws['!cols'] = NSC.map((c, i) => ({ wch: Math.min(38, Math.max(c.length + 2,
    ...rows.slice(0, 250).map(r => String(r[i] == null ? '' : r[i]).length + 2))) }));
  ws['!autofilter'] = { ref: XLSX.utils.encode_range(
    { s:{r:0,c:0}, e:{r:rows.length, c:NSC.length-1} }) };
  ws['!freeze'] = { xSplit:0, ySplit:1 };
  XLSX.utils.book_append_sheet(wb, ws, 'Not set');

  const counts = {};
  rows.forEach(r => counts[r[C_REASON]] = (counts[r[C_REASON]] || 0) + 1);
  const sum = [
    ['Catalyst 15-lb endcap — stores not set'],
    ['Selection', selLabel()],
    ['Stores in this file', rows.length],
    [],
    ['Reason', 'Stores in file', 'Stores in program', 'Have the 36 bags', 'Inbound'],
    ...NS.reasons.filter(r => nsSel.has(r.key))
      .map(r => [r.label, counts[r.label] || 0, r.n, r.received, r.inbound]),
    [],
    ['Sub-reasons given'],
    ...NS.reasons.filter(r => nsSel.has(r.key) && r.subs.length)
      .flatMap(r => [[r.label], ...r.subs.map(s => ['', s[0], s[1]])]),
    [],
    ['Program stores', DATA.n_endcap],
    ['Confirmed set', DATA.n_set],
    ['Not set (reason filed)', DATA.n_exc],
    ['Not visited yet', DATA.n_unvisited],
    [],
    ['Status source', 'Merchandiser field visits, W/E 08/08/26 (WK27 survey)'],
    ['Sales / inventory source', 'Walmart weekly reports wk' + DATA.prior_week +
      ' and wk' + DATA.week + ', Sales by Store, CATALYST15ORIG'],
    ['Endcap allocation', DATA.inv.units + ' units per store'],
    ['"36 at store"', 'on hand plus the last two weeks of sell-through'],
  ];
  const ws2 = XLSX.utils.aoa_to_sheet(sum);
  ws2['!cols'] = [{wch:32},{wch:44},{wch:18},{wch:18},{wch:12}];
  XLSX.utils.book_append_sheet(wb, ws2, 'Summary');

  XLSX.writeFile(wb, nsFilename('xlsx'));
};
renderNS();

/* ---- flag list ---- */
const T1 = DATA.flags.filter(f => f.tier === 1).length;
const FNP = DATA.flags.filter(f => f.inv !== 'received').length;
document.getElementById('flagnote').textContent =
  DATA.flags.length + ' stores were confirmed set by a merchandiser yet sold zero 15-lb units in the ' +
  'endcap week. Tier 1 (' + T1 + ') also sold nothing the week before; Tier 2 (' +
  (DATA.flags.length - T1) + ') were selling and stopped. ' + FNP +
  ' of them show no endcap product at the store, so the display is likely built from backstock or ' +
  'the receipt has not posted; the other ' + (DATA.flags.length - FNP) +
  ' have the bags on site and an endcap up but no sales — those are the ones worth a photo check.';
let ft = '<tr><th>Tier</th><th>Store #</th><th>City</th><th>State</th>' +
  '<th class="num">Units ' + wkLabel(DATA.prior_week) + '</th>' +
  '<th class="num">Units ' + wkLabel(DATA.week) + '</th><th class="num">On hand now</th>' +
  '<th>Endcap product</th></tr>';
DATA.flags.forEach(f => {
  ft += '<tr><td><span class="tier' + f.tier + '">Tier ' + f.tier + '</span></td>' +
    '<td>' + f.store + '</td><td>' + f.city + '</td><td>' + f.state + '</td>' +
    '<td class="num">' + f.q26 + '</td><td class="num">' + f.q27 + '</td>' +
    '<td class="num">' + f.oh + '</td><td>' +
    (f.inv === 'received' ? '<b>at store — check set</b>' : INV_LABELS[f.inv].toLowerCase()) +
    '</td></tr>';
});
document.getElementById('flagtable').innerHTML = ft;
document.getElementById('csvbtn').onclick = () => {
  const rows = [['tier','store_number','city','state',
    'units_wk' + DATA.prior_week,'units_wk' + DATA.week,'on_hand','endcap_product']];
  DATA.flags.forEach(f => rows.push([f.tier,f.store,f.city,f.state,f.q26,f.q27,f.oh,f.inv]));
  const blob = new Blob([rows.map(r => r.join(',')).join('\n')], {type:'text/csv'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'endcap_zero_sale_flags_wk' + DATA.week + '.csv';
  a.click();
};

/* ---- footer ---- */
document.getElementById('foot').innerHTML =
  'Definitions: U/S/W = units per store per week. &ldquo;Set&rdquo; = the merchandiser answered Yes to ' +
  '&ldquo;is the Catalyst feature product set on a feature space&rdquo; on the W/E 08/08/26 field visit ' +
  '(' + fmt(DATA.n_set) + ' stores, of which ' + (DATA.set_where['Endcap'] || 0) + ' on a true endcap). ' +
  'Every other visited store carries the reason its merchandiser filed; ' + fmt(DATA.n_unvisited) +
  ' program stores had no visit logged and are counted separately, never as set. ' +
  'Sales from Walmart weekly reports ' + wkLabel(DATA.prior_week) + ' and ' + wkLabel(DATA.week) +
  ' (Sales by Store, CATALYST15ORIG). Retail estimate at the $15.97 rollback price.';

if (localStorage.getItem(AUTH_KEY) === 'ok') showApp();
else document.getElementById('pw').focus();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
