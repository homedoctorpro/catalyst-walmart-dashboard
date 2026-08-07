"""
Endcap Rollout Map
==================
Reads the weekly "catalyst store receipts wk###### - endcaps.xlsx" workbook,
classifies each store's 36-unit endcap status, and writes a self-contained
Leaflet map -> endcap_rollout_map.html (publish via GitHub Pages).

Status logic (per-store endcap allocation is 36 units):
  received : Store Receipt QTY >= 36 (endcap arrived at the store)
  moving   : endcap still inbound, some units in transit / DC warehouse
  onorder  : endcap still inbound, all 36 still at "On Order"
  short    : total pipeline (receipt+transit+wh+onorder) < 36
Small receipts (1-13 units) are ordinary shelf replenishment, not endcap
pieces — every such store also shows a full 36 in the pipeline.
"""
import glob
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
GEO = os.path.join(HERE, "stores_geo.json")
OUTPUT = os.path.join(HERE, "endcap_rollout_map.html")
ENDCAP_UNITS = 36

STATUS = {
    "received": {"label": "Endcap received (36+)", "color": "#1a9850"},
    "moving":   {"label": "Coming — in transit / DC", "color": "#4575b4"},
    "onorder":  {"label": "Coming — still on order", "color": "#f4a340"},
    "short":    {"label": "Pipeline short of 36", "color": "#d73027"},
}


def latest_workbook():
    files = sorted(glob.glob(os.path.join(
        HERE, "catalyst store receipts wk* - endcaps.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        raise SystemExit("no 'catalyst store receipts wk* - endcaps.xlsx' found")
    return files[-1]


def classify(receipt, wh, onorder, transit):
    pipeline = transit + wh + onorder
    if receipt >= ENDCAP_UNITS:
        return "received"
    if receipt + pipeline < ENDCAP_UNITS:
        return "short"
    if transit + wh > 0:
        return "moving"
    return "onorder"


def main():
    path = latest_workbook()
    week = re.search(r"wk(\d{6})", os.path.basename(path)).group(1)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    geo = json.load(open(GEO, encoding="utf-8"))
    stores, missing = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[6] is None:
            continue
        zip5 = str(r[7] or "")[:5]
        receipt, onhand = r[8] or 0, r[9] or 0
        wh, onorder, transit = r[10] or 0, r[11] or 0, r[12] or 0
        g = geo.get(zip5)
        if not g:
            missing.append(zip5)
            continue
        stores.append({
            "store": r[6], "city": (r[4] or "").title(), "state": r[5],
            "zip": zip5, "lat": g["lat"], "lon": g["lon"],
            "receipt": receipt, "onhand": onhand, "wh": wh,
            "onorder": onorder, "transit": transit,
            "status": classify(receipt, wh, onorder, transit),
        })

    if missing:  # fetch any new zips once and persist them to the cache
        import pgeocode
        nomi = pgeocode.Nominatim("us")
        for z in sorted(set(missing)):
            rec = nomi.query_postal_code(z)
            if rec is not None and rec.latitude == rec.latitude:
                geo[z] = {"lat": round(float(rec.latitude), 4),
                          "lon": round(float(rec.longitude), 4)}
        json.dump(geo, open(GEO, "w", encoding="utf-8"))
        print(f"[ok] geocoded {len(set(missing))} new zips; re-run to include")

    counts = {k: sum(1 for s in stores if s["status"] == k) for k in STATUS}
    payload = {"week": week, "units": ENDCAP_UNITS, "stores": stores,
               "status": STATUS, "counts": counts}

    html = TEMPLATE.replace("/*DATA*/", json.dumps(payload, separators=(",", ":")))
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[ok] week {week}: {len(stores)} stores mapped "
          f"({', '.join(f'{k}={v}' for k, v in counts.items())})")
    print(f"[ok] wrote {OUTPUT}")


TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Catalyst Endcap Rollout Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  html, body { margin:0; height:100%; font-family:system-ui,Segoe UI,Arial,sans-serif; }
  #map { position:absolute; inset:0; }
  .hdr { position:absolute; top:10px; left:50px; right:10px; z-index:1000;
         background:#fff; border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.3);
         padding:8px 14px; display:inline-block; max-width:calc(100% - 80px); }
  .hdr h1 { margin:0; font-size:16px; }
  .hdr .sub { color:#666; font-size:12px; margin-top:2px; }
  .legend { position:absolute; bottom:20px; left:10px; z-index:1000; background:#fff;
            border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.3); padding:10px 14px;
            font-size:13px; }
  .legend .row { display:flex; align-items:center; gap:8px; margin:4px 0;
                 cursor:pointer; user-select:none; }
  .legend .row.off { opacity:.35; }
  .legend .dot { width:12px; height:12px; border-radius:50%; flex:0 0 auto; }
  .legend .n { color:#666; margin-left:auto; padding-left:12px; }
  .legend .hint { color:#999; font-size:11px; margin-top:6px; }
  .popup b { font-size:13px; }
  .popup table { border-collapse:collapse; margin-top:4px; font-size:12px; }
  .popup td { padding:1px 8px 1px 0; }
</style>
</head>
<body>
<div id="map"></div>
<div class="hdr">
  <h1>Catalyst 15-lb Endcap Rollout</h1>
  <div class="sub" id="sub"></div>
</div>
<div class="legend" id="legend"></div>
<script>
const DATA = /*DATA*/;
const map = L.map('map').setView([39.5, -96.5], 4);
L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',
  { attribution:'&copy; OpenStreetMap &copy; CARTO', maxZoom:18 }).addTo(map);

const total = DATA.stores.length;
const wk = DATA.week;
document.getElementById('sub').textContent =
  total + ' stores \\u00b7 ' + DATA.units + ' units per endcap \\u00b7 week ' + wk;

const groups = {};
for (const key in DATA.status) groups[key] = L.layerGroup().addTo(map);

for (const s of DATA.stores) {
  const st = DATA.status[s.status];
  const m = L.circleMarker([s.lat, s.lon], {
    radius:5, color:'#333', weight:.5, fillColor:st.color, fillOpacity:.85 });
  m.bindPopup('<div class="popup"><b>Store #' + s.store + '</b> \\u2014 ' +
    s.city + ', ' + s.state + ' ' + s.zip +
    '<br><span style="color:' + st.color + ';font-weight:600">' + st.label + '</span>' +
    '<table>' +
    '<tr><td>Received this week</td><td><b>' + s.receipt + '</b></td></tr>' +
    '<tr><td>On hand</td><td>' + s.onhand + '</td></tr>' +
    '<tr><td>In transit</td><td>' + s.transit + '</td></tr>' +
    '<tr><td>In DC warehouse</td><td>' + s.wh + '</td></tr>' +
    '<tr><td>On order</td><td>' + s.onorder + '</td></tr></table></div>');
  m.addTo(groups[s.status]);
}

const legend = document.getElementById('legend');
for (const key in DATA.status) {
  const st = DATA.status[key], n = DATA.counts[key];
  const pct = total ? Math.round(100 * n / total) : 0;
  const row = document.createElement('div');
  row.className = 'row';
  row.innerHTML = '<span class="dot" style="background:' + st.color + '"></span>' +
    st.label + '<span class="n">' + n.toLocaleString() + ' (' + pct + '%)</span>';
  row.onclick = () => {
    const off = row.classList.toggle('off');
    off ? map.removeLayer(groups[key]) : map.addLayer(groups[key]);
  };
  legend.appendChild(row);
}
const hint = document.createElement('div');
hint.className = 'hint';
hint.textContent = 'Click a row to show/hide \\u00b7 click a dot for detail';
legend.appendChild(hint);
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
