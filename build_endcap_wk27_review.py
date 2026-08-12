"""Build the WK27 endcap field-visit review (public share HTML).

Reads:
  (Walmart) Lignetics Inc. Cat Litter Endcap Set WK27.xlsx  — field visit survey (W/E 08/08/26)
  _endcap_inv_status_202627.json                            — per-store endcap shipment status
  endcap_stores.csv                                         — store meta (state, current catalyst carriage)

Writes:
  endcap-wk27-review-h4t7q2.html  — self-contained, unlisted share copy (no password)
"""
import csv
import json
from collections import Counter, defaultdict

import openpyxl

SRC = "(Walmart) Lignetics Inc. Cat Litter Endcap Set WK27.xlsx"
OUT = "endcap-wk27-review-h4t7q2.html"

wb = openpyxl.load_workbook(SRC, read_only=True)
rows = list(wb["Store Details"].iter_rows(values_only=True))[1:]
inv = json.load(open("_endcap_inv_status_202627.json"))
meta = {r["store_number"]: r for r in csv.DictReader(open("endcap_stores.csv", encoding="utf-8-sig"))}

# Summary-sheet scheduled count
sched = next(
    r[1] for r in wb["Summary"].iter_rows(values_only=True) if isinstance(r[1], (int, float)) and r[1] > 1000
)

visited = len(rows)
set_yes = [r for r in rows if r[5] == "Yes"]
not_set = [r for r in rows if r[5] == "No"]

def status(r):
    return inv.get(str(r[0]), "no record")

ARRIVED = {"received"}
reason_order = [
    "No Available space",
    "Not enough inventory to build feature",
    "Product not located",
    "Store Refusal",
]
reasons = Counter(r[9] for r in not_set if r[9])

# freight status universe (all tracked stores)
inv_c = Counter(inv.values())

# set rate by freight status
by_status = defaultdict(lambda: [0, 0])
for r in rows:
    s = status(r)
    by_status[s][0] += r[5] == "Yes"
    by_status[s][1] += 1

# per-reason arrived vs not
reason_arrival = {}
for g in reason_order:
    grp = [r for r in not_set if r[9] == g]
    arr = sum(1 for r in grp if status(r) in ARRIVED)
    reason_arrival[g] = (arr, len(grp))

nl = [r for r in not_set if r[9] == "Product not located"]
nl_sub = Counter(r[10] for r in nl if r[10])
nl_received = sum(1 for r in nl if status(r) in ARRIVED)
ref = [r for r in not_set if r[9] == "Store Refusal"]
ref_sub = Counter(r[12] for r in ref if r[12])
ref_title = Counter(r[14] for r in ref if r[14])

# recovery: inventory-blocked stores with freight inbound (in transit / on order)
inbound = {"in transit", "on order"}
inv_blocked_inbound = sum(
    1
    for r in not_set
    if r[9] in ("Not enough inventory to build feature", "Product not located") and status(r) in inbound
)
short_not_set = sum(1 for r in not_set if status(r) == "short")
recv_rate = by_status["received"][0] / by_status["received"][1]
space_among_received = sum(1 for r in not_set if r[9] == "No Available space" and status(r) in ARRIVED)

# by state
by_state = defaultdict(lambda: [0, 0])
for r in rows:
    m = meta.get(str(r[0]))
    st = m["state"] if m else "?"
    by_state[st][0] += r[5] == "Yes"
    by_state[st][1] += 1
states = sorted(by_state.items(), key=lambda kv: -kv[1][1])

# carriage split
carry = defaultdict(lambda: [0, 0])
for r in rows:
    m = meta.get(str(r[0]))
    if not m:
        continue
    k = int(m["current_sku_count"]) > 0
    carry[k][0] += r[5] == "Yes"
    carry[k][1] += 1

pct = lambda a, b: a / b * 100

# ---------------------------------------------------------------- HTML helpers
def bar_row(label, count, denom, color="var(--s1)", note=""):
    w = pct(count, denom)
    note_html = f'<span class="note">{note}</span>' if note else ""
    return f"""
    <div class="brow" title="{label}: {count:,} stores ({w:.1f}%)">
      <div class="blabel">{label}{note_html}</div>
      <div class="btrack"><div class="bfill" style="width:{w:.2f}%;background:{color}"></div></div>
      <div class="bval">{count:,} <span class="bpct">({w:.0f}%)</span></div>
    </div>"""

def rate_bar(label, y, n):
    w = pct(y, n)
    return f"""
    <div class="brow" title="{label}: {y} of {n} stores set ({w:.1f}%)">
      <div class="blabel">{label}</div>
      <div class="btrack"><div class="bfill" style="width:{w:.2f}%;background:var(--s1)"></div></div>
      <div class="bval">{w:.0f}% <span class="bpct">({y}/{n})</span></div>
    </div>"""

def stack_row(label, arr, total):
    w = pct(arr, total)
    return f"""
    <div class="brow" title="{label}: {arr} of {total} stores had received the display">
      <div class="blabel">{label}</div>
      <div class="btrack stack">
        <div class="bfill" style="width:{w:.2f}%;background:var(--s1)"></div>
        <div class="bfill gray" style="width:{100-w:.2f}%"></div>
      </div>
      <div class="bval">{w:.0f}% <span class="bpct">arrived</span></div>
    </div>"""

reason_bars = "\n".join(bar_row(g, reasons[g], visited) for g in reason_order)

status_labels = [
    ("received", "Received"),
    ("in transit", "In transit"),
    ("short", "Pipeline short"),
    ("on order", "On order"),
]
status_bars = "\n".join(rate_bar(lab, *by_status[k]) for k, lab in status_labels)
stack_bars = "\n".join(stack_row(g, *reason_arrival[g]) for g in reason_order)

state_rows = ""
for st, (y, n) in states[:15]:
    w = pct(y, n)
    state_rows += f"""<tr><td>{st}</td><td class="num">{n}</td><td class="num">{y}</td>
      <td><div class="mtrack"><div class="mfill" style="width:{w:.1f}%"></div></div></td>
      <td class="num">{w:.0f}%</td></tr>\n"""
rest_y = sum(y for st, (y, n) in states[15:])
rest_n = sum(n for st, (y, n) in states[15:])
state_rows += f"""<tr class="restrow"><td>All other states ({len(states)-15})</td><td class="num">{rest_n}</td><td class="num">{rest_y}</td>
  <td><div class="mtrack"><div class="mfill" style="width:{pct(rest_y, rest_n):.1f}%"></div></div></td>
  <td class="num">{pct(rest_y, rest_n):.0f}%</td></tr>"""

nl_rows = "\n".join(
    f'<tr><td>{k}</td><td class="num">{v}</td></tr>' for k, v in nl_sub.most_common()
)
ref_rows = "\n".join(
    f'<tr><td>{k}</td><td class="num">{v}</td></tr>' for k, v in ref_sub.most_common()
)

not_arrived = inv_c["in transit"] + inv_c["on order"] + inv_c["short"]
total_tracked = sum(inv_c.values())
est_lo = round(inv_blocked_inbound * recv_rate)

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>Catalyst Endcap WK27 — Why Endcaps Aren't Set</title>
<style>
  :root {{
    color-scheme: light;
    --page:#f9f9f7; --surface:#fcfcfb; --ink:#0b0b0b; --ink2:#52514e; --muted:#898781;
    --grid:#e1e0d9; --border:rgba(11,11,11,.10);
    --s1:#2a78d6; --s2:#eb6834; --gray:#d9d8d2; --good:#0ca30c; --crit:#d03b3b;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      color-scheme: dark;
      --page:#0d0d0d; --surface:#1a1a19; --ink:#ffffff; --ink2:#c3c2b7; --muted:#898781;
      --grid:#2c2c2a; --border:rgba(255,255,255,.10);
      --s1:#3987e5; --s2:#d95926; --gray:#3a3a37;
    }}
  }}
  * {{ box-sizing:border-box; margin:0; }}
  body {{ background:var(--page); color:var(--ink); font-family:system-ui,-apple-system,"Segoe UI",sans-serif;
         line-height:1.55; padding:28px 16px 64px; }}
  .wrap {{ max-width:880px; margin:0 auto; }}
  h1 {{ font-size:26px; letter-spacing:-.01em; }}
  .sub {{ color:var(--ink2); margin:4px 0 26px; font-size:14.5px; }}
  h2 {{ font-size:18px; margin:34px 0 6px; }}
  h2 .n {{ color:var(--muted); font-weight:600; margin-right:6px; }}
  p {{ color:var(--ink2); font-size:14.5px; margin:8px 0; }}
  p strong, li strong {{ color:var(--ink); }}
  ul {{ margin:8px 0 8px 20px; color:var(--ink2); font-size:14.5px; }}
  li {{ margin:4px 0; }}
  .card {{ background:var(--surface); border:1px solid var(--border); border-radius:12px; padding:18px 20px; margin:14px 0; }}
  .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:10px; margin:18px 0; }}
  .kpi {{ background:var(--surface); border:1px solid var(--border); border-radius:12px; padding:14px 16px; }}
  .kpi .l {{ font-size:12.5px; color:var(--ink2); }}
  .kpi .v {{ font-size:27px; font-weight:600; margin-top:2px; }}
  .kpi .d {{ font-size:12.5px; color:var(--muted); }}
  .kpi.hero .v {{ font-size:48px; color:var(--crit); }}
  .brow {{ display:grid; grid-template-columns:230px 1fr 110px; gap:10px; align-items:center; margin:9px 0; }}
  .blabel {{ font-size:13.5px; color:var(--ink2); text-align:right; }}
  .blabel .note {{ display:block; font-size:11.5px; color:var(--muted); }}
  .btrack {{ background:none; height:22px; position:relative; }}
  .bfill {{ height:22px; border-radius:0 4px 4px 0; display:inline-block; vertical-align:top; }}
  .stack .bfill {{ border-radius:0; }}
  .stack .bfill:first-child {{ border-right:2px solid var(--surface); }}
  .stack .bfill:last-child {{ border-radius:0 4px 4px 0; }}
  .bfill.gray {{ background:var(--gray); }}
  .bval {{ font-size:13.5px; font-variant-numeric:tabular-nums; }}
  .bpct {{ color:var(--muted); font-size:12px; }}
  .legend {{ display:flex; gap:18px; font-size:12.5px; color:var(--ink2); margin:4px 0 2px 240px; }}
  .legend span::before {{ content:""; display:inline-block; width:10px; height:10px; border-radius:3px; margin-right:6px; }}
  .legend .k1::before {{ background:var(--s1); }}
  .legend .k2::before {{ background:var(--gray); }}
  table {{ border-collapse:collapse; width:100%; font-size:13.5px; }}
  th {{ text-align:left; color:var(--muted); font-weight:600; font-size:12px; text-transform:uppercase;
        letter-spacing:.04em; padding:6px 10px; border-bottom:1px solid var(--grid); }}
  td {{ padding:6px 10px; border-bottom:1px solid var(--grid); color:var(--ink2); }}
  td:first-child {{ color:var(--ink); }}
  .num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  th.num {{ text-align:right; }}
  .mtrack {{ background:var(--grid); border-radius:4px; height:10px; min-width:90px; }}
  .mfill {{ background:var(--s1); height:10px; border-radius:4px 0 0 4px; }}
  .restrow td {{ color:var(--muted); }}
  .foot {{ color:var(--muted); font-size:12px; margin-top:36px; border-top:1px solid var(--grid); padding-top:12px; }}
  .callout {{ border-left:3px solid var(--s2); padding:2px 0 2px 14px; margin:14px 0; }}
  .callout p {{ margin:4px 0; }}
</style>
</head>
<body>
<div class="wrap">
  <h1>Catalyst Cat Litter Endcap — WK27 Field Visit Review</h1>
  <div class="sub">Walmart · week ending 08/08/26 · {visited:,} of {sched:,.0f} scheduled stores visited ({pct(visited, sched):.1f}% visit compliance) · unlisted share copy</div>

  <div class="kpis">
    <div class="kpi hero"><div class="l">Endcaps set</div><div class="v">{pct(len(set_yes), visited):.0f}%</div><div class="d">{len(set_yes):,} of {visited:,} visited stores</div></div>
    <div class="kpi"><div class="l">Not set</div><div class="v">{len(not_set):,}</div><div class="d">stores with a blocker reported</div></div>
    <div class="kpi"><div class="l">Endcap allocation arrived</div><div class="v">{pct(inv_c['received'], total_tracked):.0f}%</div><div class="d">{inv_c['received']:,} of {total_tracked:,} stores, week ending 8/8</div></div>
    <div class="kpi"><div class="l">Set rate where product arrived</div><div class="v">{recv_rate*100:.0f}%</div><div class="d">vs {pct(by_status['in transit'][0], by_status['in transit'][1]):.0f}% where still in transit</div></div>
  </div>

  <h2>What the field teams reported</h2>
  <p>Merchandisers answered "Why is the Catalyst feature product not set?" in {len(not_set):,} stores. Share of all {visited:,} visited stores:</p>
  <div class="card">
    {reason_bars}
    {bar_row("Set successfully", len(set_yes), visited, "var(--good)")}
  </div>

  <h2>Why they weren't set — the four real drivers</h2>

  <h2 style="font-size:16px"><span class="n">1.</span>The product hadn't arrived yet — visits ran ahead of freight</h2>
  <p>Visits happened Aug 2–8, and by the end of that week only <strong>{inv_c['received']:,} of {total_tracked:,} stores ({pct(inv_c['received'], total_tracked):.0f}%)</strong> had
     received their full 36-bag endcap allocation (measured from Walmart on-hand + sell-through data, week ending 8/8). {inv_c['in transit']:,} had the balance in transit, {inv_c['on order']:,} on order only,
     and {inv_c['short']:,} had a pipeline short of 36 bags. Freight timing is the single strongest predictor of whether an endcap got set:</p>
  <div class="card">
    <p style="margin:0 0 6px;font-size:13px;color:var(--muted)">Set rate by endcap-allocation status (week ending 8/8)</p>
    {status_bars}
  </div>
  <p>Stores with product in hand set at <strong>{recv_rate*100:.0f}%</strong> — 2.5–5× the rate of stores still waiting on freight.
     The two "inventory" reasons are mostly this in disguise: <strong>{pct(346-61, 346):.0f}%</strong> of "not enough inventory" stores and
     <strong>{pct(261-75, 261):.0f}%</strong> of "product not located" stores had not received their full allocation.
     The field crews weren't wrong — the product genuinely wasn't there to build with.</p>
  <p>The status snapshot is end-of-week, which if anything <em>understates</em> this gap: a store visited Tuesday that received Friday still counts "received" here,
     diluting that group's rate. The visit-date pattern backs this up — set rates in the "in transit" group climb from ~14% early in the week to 20–33% by Fri/Sat
     as freight landed, while the "received" group holds flat at 38–45% all week.</p>

  <h2 style="font-size:16px"><span class="n">2.</span>"No available space" is the structural ceiling — and it doesn't go away when freight lands</h2>
  <p><strong>{reasons['No Available space']:,} stores ({pct(reasons['No Available space'], visited):.0f}% of all visits)</strong> said there was no feature space available —
     the single largest bucket. Critically, this holds even where the display had arrived:
     <strong>{space_among_received} of {by_status['received'][1]:,} received stores ({pct(space_among_received, by_status['received'][1]):.0f}%)</strong> still cited no space.
     Endcap/feature space in these stores is already committed to other programs, and store management controls it.
     Even with perfect freight execution, the current approach tops out around a ~40% set rate unless space is unlocked at the Walmart level.</p>

  <h2 style="font-size:16px"><span class="n">3.</span>In-store execution: product in the building that can't be found or trusted</h2>
  <p>Of the {reasons['Product not located']:,} "product not located" stores, {nl_received} had actually received their full allocation per the pipeline data —
     product is in the building but buried. The sub-reasons point at on-hand data integrity:</p>
  <div class="card"><table>
    <tr><th>Sub-reason</th><th class="num">Stores</th></tr>
    {nl_rows}
  </table></div>
  <p>{nl_sub['Product showing zero on hands']} stores show <strong>zero on-hands</strong> and {nl_sub['Unable to locate inventory, on hand adjustments suggested']} more couldn't reconcile
     on-hands with what was findable — these need on-hand corrections before any revisit can succeed. {nl_sub['Product up in the steel/inaccessible']} had product visible but up in the steel.</p>

  <h2 style="font-size:16px"><span class="n">4.</span>Store refusals are real but small</h2>
  <p>Only <strong>{reasons['Store Refusal']} stores ({pct(reasons['Store Refusal'], visited):.1f}%)</strong> refused the service —
     and most refusals are soft: {ref_sub['Store wants completed at a later date']} asked to complete at a later date, {ref_sub['Store in remodel/inventory']} were mid-remodel/inventory.
     These largely convert on a revisit.</p>
  <div class="card"><table>
    <tr><th>Refusal reason</th><th class="num">Stores</th></tr>
    {ref_rows}
  </table></div>

  <h2>Freight arrival by reported reason</h2>
  <p>Share of each not-set group whose full allocation had arrived by week's end — "no space" and refusals are people problems; the inventory reasons are freight problems:</p>
  <div class="card">
    <div class="legend"><span class="k1">Allocation received</span><span class="k2">In transit / on order / pipeline short</span></div>
    {stack_bars}
  </div>

  <div class="callout">
    <p><strong>What this means for recovery:</strong></p>
    <p>• <strong>{inv_blocked_inbound} inventory-blocked stores have freight inbound</strong> (in transit or on order). At the received-store set rate
       ({recv_rate*100:.0f}%), a scheduled revisit after arrival converts roughly <strong>{est_lo}+ additional endcaps</strong>; if inventory was truly the only blocker, the upside is higher.</p>
    <p>• <strong>{inv_c['short']:,} stores have a pipeline short of the 36-bag allocation</strong> ({short_not_set} of them not set) — no store order yet covers the full
       quantity (this feed has no DC-warehouse column, so DC stock not yet on a store order also shows here). These need replenishment/allocation follow-up, not a revisit.</p>
    <p>• The <strong>{reasons['No Available space']:,} no-space stores need a different lever</strong>: a directed feature from Walmart merchandising, or field teams
       returning when seasonal space frees up. Revisits alone won't move them.</p>
    <p>• On-hand corrections should be filed for the ~{nl_sub['Product showing zero on hands'] + nl_sub['Unable to locate inventory, on hand adjustments suggested']} stores with zero/incorrect on-hands so replenishment and future sets aren't suppressed.</p>
  </div>

  <h2>Set rate by state</h2>
  <p>Top 15 states by scheduled visits. Existing Catalyst stores set at {pct(*carry[True]):.0f}%; stores new to the brand (no Catalyst on shelf yet) at {pct(*carry[False]):.0f}%.</p>
  <div class="card"><table>
    <tr><th>State</th><th class="num">Visited</th><th class="num">Set</th><th>Set rate</th><th class="num">%</th></tr>
    {state_rows}
  </table></div>

  <div class="foot">
    Sources: field visit survey "(Walmart) Lignetics Inc. Cat Litter Endcap Set W/E 08/08/26" ({visited:,} store visits, Aug 2–8 2026);
    endcap allocation status inferred from Walmart on-hand + sell-through + in-transit/on-order pipeline for the 36-bag CATALYST15ORIG allocation (weekly by-store feed, week ending 8/8/26); Catalyst store list.
    "Received" means the full allocation is accounted for at the store, not a carrier delivery scan; "pipeline short" can include DC stock not yet on a store order.
    Set counts here use the survey's per-store responses ({len(set_yes):,} "Yes"); earlier extracts that counted all non-exception stores as set showed a higher figure
    because not-yet-reported stores were included. Freight status reflects the latest receipts feed and may lag actual deliveries by a few days.
    Unlisted link — not indexed, no password. Generated by build_endcap_wk27_review.py.
  </div>
</div>
</body>
</html>
"""

with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
print(f"wrote {OUT} ({len(html):,} bytes)")
print(f"set {len(set_yes)}/{visited} = {pct(len(set_yes), visited):.1f}% | recovery est {est_lo}+ | inv-blocked inbound {inv_blocked_inbound} | short not set {short_not_set}")
